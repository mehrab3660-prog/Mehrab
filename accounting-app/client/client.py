# -*- coding: utf-8 -*-
"""
کلاینت دسکتاپ حسابداری مغازه الکتریکی
این برنامه روی هر کامپیوتر (سرور یا کلاینت‌های دیگر) نصب و اجرا میشه
و از طریق شبکه داخلی (LAN) به سرور وصل میشه.
"""
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import requests
import json
import os
import subprocess
import sys
import tempfile
import pdf_generator
import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl
from openpyxl.utils import get_column_letter

# تلاش برای استفاده از ttkbootstrap برای ظاهر گرافیکی مدرن‌تر؛
# اگر نصب نبود، برنامه با ظاهر استاندارد tkinter بدون خطا اجرا می‌شود.
try:
    import ttkbootstrap as ttkb
    _HAS_TTKBOOTSTRAP = True
except ImportError:
    _HAS_TTKBOOTSTRAP = False

_BASE_APP = ttkb.Window if _HAS_TTKBOOTSTRAP else tk.Tk

CONFIG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")


def load_server_url():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f).get("server_url", "http://127.0.0.1:5050")
    return "http://127.0.0.1:5050"


def save_server_url(url):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump({"server_url": url}, f)


class App(_BASE_APP):
    def __init__(self):
        if _HAS_TTKBOOTSTRAP:
            super().__init__(themename="flatly")
        else:
            super().__init__()
        self.title("حسابداری مغازه الکتریکی")
        self.geometry("1000x650")
        self.server_url = load_server_url()
        self.user = None
        self.token = None
        self.dark_mode = False
        self.show_login()

    # ---------------- ورود ----------------
    def show_login(self):
        for w in self.winfo_children():
            w.destroy()

        frame = ttk.Frame(self, padding=40)
        frame.pack(expand=True)

        ttk.Label(frame, text="ورود به سیستم حسابداری", font=("Tahoma", 16, "bold")).grid(
            row=0, column=0, columnspan=2, pady=20)

        ttk.Label(frame, text="آدرس سرور:").grid(row=1, column=1, sticky="e", pady=5)
        server_entry = ttk.Entry(frame, width=30, justify="left")
        server_entry.insert(0, self.server_url)
        server_entry.grid(row=1, column=0, pady=5)

        ttk.Label(frame, text="نام کاربری:").grid(row=2, column=1, sticky="e", pady=5)
        user_entry = ttk.Entry(frame, width=30, justify="left")
        user_entry.insert(0, "admin")
        user_entry.grid(row=2, column=0, pady=5)

        ttk.Label(frame, text="رمز عبور:").grid(row=3, column=1, sticky="e", pady=5)
        pass_entry = ttk.Entry(frame, width=30, show="*", justify="left")
        pass_entry.grid(row=3, column=0, pady=5)

        captcha_state = {"id": None}
        captcha_label_var = tk.StringVar(value="...")

        ttk.Label(frame, text="کد امنیتی:").grid(row=4, column=1, sticky="e", pady=5)
        captcha_row = ttk.Frame(frame)
        captcha_row.grid(row=4, column=0, pady=5, sticky="w")
        ttk.Label(captcha_row, textvariable=captcha_label_var, font=("Tahoma", 11, "bold")).pack(side="right", padx=(0, 6))
        captcha_entry = ttk.Entry(captcha_row, width=10, justify="left")
        captcha_entry.pack(side="right")

        def refresh_captcha():
            server = server_entry.get().strip()
            try:
                r = requests.get(f"{server}/captcha", timeout=5)
                data = r.json()
                captcha_state["id"] = data.get("captcha_id")
                captcha_label_var.set(data.get("question", "?"))
                captcha_entry.delete(0, tk.END)
            except requests.exceptions.RequestException:
                captcha_label_var.set("خطا در دریافت کد")

        ttk.Button(captcha_row, text="🔄", width=3, command=refresh_captcha).pack(side="right", padx=(6, 0))

        def do_login():
            self.server_url = server_entry.get().strip()
            save_server_url(self.server_url)
            try:
                r = requests.post(f"{self.server_url}/login", json={
                    "username": user_entry.get(),
                    "password": pass_entry.get(),
                    "captcha_id": captcha_state["id"],
                    "captcha_answer": captcha_entry.get(),
                }, timeout=5)
                data = r.json()
                if data.get("ok"):
                    self.user = data["user"]
                    self.token = data.get("token")
                    self.maybe_force_password_change(self.show_main)
                else:
                    messagebox.showerror("خطا", data.get("message", "ورود ناموفق بود"))
                    refresh_captcha()
            except requests.exceptions.RequestException:
                messagebox.showerror("خطا در اتصال",
                                      "اتصال به سرور برقرار نشد.\nمطمئن شوید سرور روشن است و آدرس آن درست وارد شده.")

        ttk.Button(frame, text="ورود", command=do_login).grid(row=5, column=0, columnspan=2, pady=20)
        refresh_captcha()

    # ---------------- صفحه اصلی ----------------
    def show_main(self):
        for w in self.winfo_children():
            w.destroy()

        top_bar = ttk.Frame(self, padding=5)
        top_bar.pack(fill="x")
        ttk.Label(top_bar, text=f'کاربر: {self.user["username"]} ({"مدیر" if self.user.get("role")=="admin" else "کارمند"})',
                  font=("Tahoma", 10)).pack(side="right", padx=10)
        ttk.Button(top_bar, text="تغییر رمز عبور من", command=self.open_change_own_password_dialog).pack(side="right", padx=10)
        ttk.Button(top_bar, text="تغییر حالت تاریک/روشن", command=self.toggle_dark_mode).pack(side="left", padx=10)

        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True)
        self.notebook = notebook
        self.tab_refresh_callbacks = {}

        def on_tab_changed(event):
            try:
                current_frame_name = notebook.select()
                current_widget = self.nametowidget(current_frame_name)
                cb = self.tab_refresh_callbacks.get(current_widget)
                if cb:
                    cb()
            except Exception:
                pass

        notebook.bind("<<NotebookTabChanged>>", on_tab_changed)

        self.tab_dashboard = ttk.Frame(notebook)
        self.tab_items = ttk.Frame(notebook)
        self.tab_sale = ttk.Frame(notebook)
        self.tab_purchase = ttk.Frame(notebook)
        self.tab_parties = ttk.Frame(notebook)
        self.tab_cash = ttk.Frame(notebook)
        self.tab_checks = ttk.Frame(notebook)
        self.tab_history = ttk.Frame(notebook)
        self.tab_monthly = ttk.Frame(notebook)
        self.tab_extra_reports = ttk.Frame(notebook)

        notebook.add(self.tab_dashboard, text="خلاصه وضعیت")
        notebook.add(self.tab_items, text="کالاها")
        notebook.add(self.tab_sale, text="فاکتور فروش")
        notebook.add(self.tab_purchase, text="فاکتور خرید")
        notebook.add(self.tab_history, text="تاریخچه و چاپ فاکتورها")
        notebook.add(self.tab_parties, text="مشتریان و تامین‌کنندگان")
        notebook.add(self.tab_checks, text="چک‌ها")
        notebook.add(self.tab_cash, text="صندوق")
        notebook.add(self.tab_monthly, text="گزارش ماهانه")
        notebook.add(self.tab_extra_reports, text="بدهکاران/بستانکاران/پرفروش‌ها")

        self.build_dashboard(self.tab_dashboard)
        self.build_items_tab(self.tab_items)
        self.build_invoice_tab(self.tab_sale, "sale")
        self.build_invoice_tab(self.tab_purchase, "purchase")
        self.build_history_tab(self.tab_history)
        self.build_parties_tab(self.tab_parties)
        self.build_checks_tab(self.tab_checks)
        self.build_cash_tab(self.tab_cash)
        self.build_monthly_tab(self.tab_monthly)
        self.build_extra_reports_tab(self.tab_extra_reports)

        # این تب‌ها فقط برای مدیر (admin) نمایش داده می‌شود
        if self.user and self.user.get("role") == "admin":
            self.tab_backup = ttk.Frame(notebook)
            self.tab_users = ttk.Frame(notebook)
            self.tab_activity = ttk.Frame(notebook)
            notebook.add(self.tab_backup, text="پشتیبان‌گیری")
            notebook.add(self.tab_users, text="مدیریت کاربران")
            notebook.add(self.tab_activity, text="لاگ کاربران")
            self.build_backup_tab(self.tab_backup)
            self.build_users_tab(self.tab_users)
            self.build_activity_log_tab(self.tab_activity)

    def toggle_dark_mode(self):
        self.dark_mode = not getattr(self, "dark_mode", False)
        if _HAS_TTKBOOTSTRAP:
            self.style.theme_use("darkly" if self.dark_mode else "flatly")
            return
        style = ttk.Style(self)
        if self.dark_mode:
            style.theme_use("clam")
            bg, fg, field_bg = "#1e1e1e", "#f0f0f0", "#2d2d2d"
            style.configure(".", background=bg, foreground=fg, fieldbackground=field_bg)
            style.configure("Treeview", background=field_bg, foreground=fg, fieldbackground=field_bg)
            style.map("Treeview", background=[("selected", "#3a6ea5")])
            self.configure(bg=bg)
        else:
            style.theme_use("vista" if sys.platform.startswith("win") else "default")
            self.configure(bg="SystemButtonFace")

    def api(self, method, path, **kwargs):
        if self.token:
            headers = kwargs.pop("headers", {}) or {}
            headers.setdefault("Authorization", f"Bearer {self.token}")
            kwargs["headers"] = headers
        try:
            r = requests.request(method, f"{self.server_url}{path}", timeout=8, **kwargs)
            if r.status_code == 401:
                messagebox.showerror("خطا", "نشست شما منقضی شده، دوباره وارد شوید")
                self.user = None
                self.token = None
                self.show_login()
                return None
            return r.json()
        except requests.exceptions.RequestException:
            messagebox.showerror("خطا", "ارتباط با سرور برقرار نشد. آیا سرور روشن است؟")
            return None

    def maybe_force_password_change(self, on_done):
        """اگه رمز عبور کاربر موقتی/اولیه باشه، قبل از ادامه یه پنجره‌ی اجباری برای انتخاب
        رمز جدید نشون می‌ده (قابل بستن نیست)؛ در پایان on_done() صدا زده می‌شه."""
        if not (self.user and self.user.get("must_change_password")):
            on_done()
            return

        win = tk.Toplevel(self)
        win.title("تغییر رمز عبور لازم است")
        win.protocol("WM_DELETE_WINDOW", lambda: None)
        win.grab_set()

        ttk.Label(win, text="برای امنیت بیشتر، قبل از ادامه باید یک رمز عبور جدید برای خودتان انتخاب کنید.",
                  wraplength=320, justify="center").pack(padx=20, pady=(20, 10))
        ttk.Label(win, text="رمز عبور جدید:").pack(padx=20, anchor="e")
        p1_entry = ttk.Entry(win, show="*", width=30)
        p1_entry.pack(padx=20, pady=5)
        ttk.Label(win, text="تکرار رمز عبور جدید:").pack(padx=20, anchor="e")
        p2_entry = ttk.Entry(win, show="*", width=30)
        p2_entry.pack(padx=20, pady=5)

        def save():
            p1, p2 = p1_entry.get(), p2_entry.get()
            if not p1 or len(p1) < 4:
                messagebox.showerror("خطا", "رمز عبور باید حداقل ۴ کاراکتر باشد")
                return
            if p1 != p2:
                messagebox.showerror("خطا", "رمز عبور و تکرار آن یکسان نیستند")
                return
            res = self.api("POST", "/users/me/password", json={"new_password": p1})
            if res and res.get("ok"):
                self.user["must_change_password"] = False
                win.grab_release()
                win.destroy()
                on_done()
            else:
                messagebox.showerror("خطا", (res or {}).get("message", "تغییر رمز عبور ناموفق بود"))

        ttk.Button(win, text="ذخیره و ادامه", command=save).pack(pady=15)
        win.bind("<Return>", lambda e: save())

    def open_change_own_password_dialog(self):
        """تغییر داوطلبانه‌ی رمز عبور خودِ کاربر (برخلاف maybe_force_password_change، این پنجره قابل بستن است)"""
        win = tk.Toplevel(self)
        win.title("تغییر رمز عبور من")

        ttk.Label(win, text="رمز عبور جدید:").pack(padx=20, pady=(20, 0), anchor="e")
        p1_entry = ttk.Entry(win, show="*", width=30)
        p1_entry.pack(padx=20, pady=5)
        ttk.Label(win, text="تکرار رمز عبور جدید:").pack(padx=20, anchor="e")
        p2_entry = ttk.Entry(win, show="*", width=30)
        p2_entry.pack(padx=20, pady=5)

        def save():
            p1, p2 = p1_entry.get(), p2_entry.get()
            if not p1 or len(p1) < 4:
                messagebox.showerror("خطا", "رمز عبور باید حداقل ۴ کاراکتر باشد")
                return
            if p1 != p2:
                messagebox.showerror("خطا", "رمز عبور و تکرار آن یکسان نیستند")
                return
            res = self.api("POST", "/users/me/password", json={"new_password": p1})
            if res and res.get("ok"):
                messagebox.showinfo("موفق", "رمز عبور با موفقیت تغییر کرد")
                win.destroy()
            else:
                messagebox.showerror("خطا", (res or {}).get("message", "تغییر رمز عبور ناموفق بود"))

        ttk.Button(win, text="ذخیره", command=save).pack(pady=15)
        win.bind("<Return>", lambda e: save())

    def export_to_excel(self, headers, rows, filename_prefix):
        """ساخت فایل اکسل از یک لیست هدر و ردیف‌ها و باز کردن آن"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(headers)
            for row in rows:
                ws.append(row)
            for i, h in enumerate(headers, start=1):
                ws.column_dimensions[get_column_letter(i)].width = max(14, len(str(h)) + 4)
            out_path = os.path.join(tempfile.gettempdir(), f'{filename_prefix}_{tempfile.gettempprefix()}.xlsx')
            wb.save(out_path)
        except Exception as e:
            messagebox.showerror("خطا در ساخت اکسل", str(e))
            return
        try:
            if sys.platform.startswith("win"):
                os.startfile(out_path)
            elif sys.platform == "darwin":
                subprocess.call(["open", out_path])
            else:
                subprocess.call(["xdg-open", out_path])
        except Exception:
            messagebox.showinfo("فایل ذخیره شد", f"فایل اکسل در این مسیر ذخیره شد:\n{out_path}")

    def print_invoice_pdf(self, invoice, items, party_name=None):
        """پرسیدن فرمت کاغذ، تولید فایل PDF فاکتور، و چاپ مستقیم یا باز کردن آن"""
        page_format = self.ask_page_format()
        if not page_format:
            return
        try:
            out_path = os.path.join(tempfile.gettempdir(), f'invoice_{invoice["id"]}.pdf')
            pdf_generator.generate_invoice_pdf(out_path, invoice, items, party_name=party_name, page_format=page_format)
        except Exception as e:
            messagebox.showerror("خطا در تولید PDF", str(e))
            return

        want_direct_print = messagebox.askyesno(
            "چاپ فاکتور",
            "فایل فاکتور ساخته شد.\n\nآیا می‌خواهید مستقیم به چاپگر پیش‌فرض ارسال شود؟\n"
            "(اگر «خیر» را بزنید، فقط فایل باز می‌شود تا خودتان چاپ کنید)"
        )
        if want_direct_print:
            ok, msg = pdf_generator.print_pdf_direct(out_path)
            if ok:
                messagebox.showinfo("ارسال به چاپگر", msg)
                return
            else:
                messagebox.showwarning("چاپ مستقیم ممکن نشد",
                                        f'{msg}\n\nفایل به‌جای آن باز می‌شود تا از همان‌جا چاپ کنید.')

        try:
            if sys.platform.startswith("win"):
                os.startfile(out_path)
            elif sys.platform == "darwin":
                subprocess.call(["open", out_path])
            else:
                subprocess.call(["xdg-open", out_path])
        except Exception:
            messagebox.showinfo("فایل ذخیره شد", f"فایل PDF در این مسیر ذخیره شد:\n{out_path}")

    def ask_page_format(self):
        """پنجره کوچک برای انتخاب سایز کاغذ چاپ فاکتور"""
        result = {"value": None}
        win = tk.Toplevel(self)
        win.title("انتخاب فرمت چاپ")
        win.geometry("320x220")
        win.grab_set()

        ttk.Label(win, text="فرمت کاغذ چاپ فاکتور را انتخاب کنید:", font=("Tahoma", 11)).pack(pady=15)
        choice = tk.StringVar(value="A5")
        options = [("A4 (کاغذ اداری کامل)", "A4"),
                   ("A5 (نصف A4 - پیش‌فرض)", "A5"),
                   ("رسید حرارتی ۸۰ میلی‌متری", "thermal80"),
                   ("رسید حرارتی ۵۸ میلی‌متری", "thermal58")]
        for label, val in options:
            ttk.Radiobutton(win, text=label, variable=choice, value=val).pack(anchor="w", padx=30, pady=3)

        def confirm():
            result["value"] = choice.get()
            win.destroy()

        ttk.Button(win, text="تأیید و ساخت PDF", command=confirm).pack(pady=15)
        self.wait_window(win)
        return result["value"]

    # ---------------- خلاصه وضعیت ----------------
    def build_dashboard(self, parent):
        frame = ttk.Frame(parent, padding=20)
        frame.pack(fill="both", expand=True)
        self.dashboard_frame = frame
        self.refresh_dashboard()
        ttk.Button(parent, text="بروزرسانی", command=self.refresh_dashboard).pack(pady=5)
        self.tab_refresh_callbacks[parent] = self.refresh_dashboard

    def refresh_dashboard(self):
        for w in self.dashboard_frame.winfo_children():
            w.destroy()
        data = self.api("GET", "/reports/summary")
        if not data:
            return
        rows = [
            ("مجموع فروش", f'{data["total_sales"]:,.0f} تومان'),
            ("مجموع خرید", f'{data["total_purchases"]:,.0f} تومان'),
            ("سود تقریبی", f'{data["estimated_profit"]:,.0f} تومان'),
            ("مجموع طلب از مشتریان", f'{data["total_debtors"]:,.0f} تومان'),
        ]
        for i, (label, val) in enumerate(rows):
            ttk.Label(self.dashboard_frame, text=label, font=("Tahoma", 12, "bold")).grid(row=i, column=1, sticky="e", pady=8, padx=10)
            ttk.Label(self.dashboard_frame, text=val, font=("Tahoma", 12)).grid(row=i, column=0, sticky="w", pady=8, padx=10)

        if data["low_stock_items"]:
            ttk.Label(self.dashboard_frame, text="⚠ کالاهای رو به اتمام:", font=("Tahoma", 12, "bold"), foreground="red").grid(
                row=len(rows) + 1, column=0, columnspan=2, sticky="w", pady=(20, 5))
            for j, item in enumerate(data["low_stock_items"]):
                ttk.Label(self.dashboard_frame, text=f'{item["name"]} - موجودی: {item["stock_qty"]}').grid(
                    row=len(rows) + 2 + j, column=0, columnspan=2, sticky="w")

    # ---------------- کالاها ----------------
    def build_items_tab(self, parent):
        ttk.Label(parent, text="اینجا کالاهایی که می‌فروشی یا نگه می‌داری تعریف می‌کنی — مثلاً «لامپ ال‌ای‌دی ۱۲ وات، برند فیلیپس».",
                  foreground="gray", padding=(10, 8, 10, 0)).pack(anchor="w")
        cat_frame = ttk.Frame(parent, padding=(10, 10, 10, 0))
        cat_frame.pack(fill="x")
        ttk.Label(cat_frame, text="دسته‌بندی کالا:").pack(side="right", padx=5)
        cat_combo_new = ttk.Combobox(cat_frame, width=15, state="readonly")
        cat_combo_new.pack(side="right", padx=5)
        self.item_category_combo = cat_combo_new

        def add_new_category():
            name = simpledialog.askstring("دسته‌بندی جدید", "نام دسته‌بندی جدید (مثلاً: روشنایی، سیم و کابل):")
            if name:
                res = self.api("POST", "/categories", json={"name": name})
                if res:
                    if res.get("ok"):
                        refresh_categories()
                    else:
                        messagebox.showerror("خطا", res.get("message", "خطا"))

        ttk.Button(cat_frame, text="افزودن دسته جدید", command=add_new_category).pack(side="right", padx=5)

        def refresh_categories():
            cats = self.api("GET", "/categories") or []
            self.categories_list = cats
            cat_combo_new["values"] = [c["name"] for c in cats]

        refresh_categories()

        top = ttk.Frame(parent, padding=10)
        top.pack(fill="x")

        labels = ["کد/بارکد", "نام کالا", "برند", "واحد", "قیمت خرید", "قیمت فروش", "موجودی", "حداقل موجودی"]
        self.item_entries = {}
        for i, lbl in enumerate(labels):
            ttk.Label(top, text=lbl).grid(row=0, column=len(labels) - 1 - i, padx=3)
            e = ttk.Entry(top, width=12)
            e.grid(row=1, column=len(labels) - 1 - i, padx=3)
            self.item_entries[lbl] = e

        ttk.Button(top, text="افزودن کالا", command=self.add_item).grid(row=1, column=len(labels), padx=10)

        cols = ("id", "code", "name", "brand", "unit", "purchase_price", "sale_price", "stock_qty", "min_stock")
        self.items_tree = ttk.Treeview(parent, columns=cols, show="headings")
        headers = {"id": "ردیف", "code": "کد/بارکد", "name": "نام", "brand": "برند", "unit": "واحد",
                   "purchase_price": "قیمت خرید", "sale_price": "قیمت فروش",
                   "stock_qty": "موجودی", "min_stock": "حداقل موجودی"}
        for c in cols:
            self.items_tree.heading(c, text=headers[c])
            self.items_tree.column(c, width=95, anchor="center")
        self.items_tree.pack(fill="both", expand=True, padx=10, pady=10)

        def show_stock_ledger():
            sel = self.items_tree.selection()
            if not sel:
                messagebox.showerror("خطا", "یک کالا را از لیست انتخاب کنید")
                return
            item_id = self.items_tree.item(sel[0])["values"][0]
            data = self.api("GET", f"/reports/stock-ledger/{item_id}")
            if not data:
                return
            win = tk.Toplevel(self)
            win.title(f'گردش انبار: {data["item"]["name"]}')
            win.geometry("650x400")
            ltree = ttk.Treeview(win, columns=("date", "type", "number", "qty", "balance"), show="headings")
            for c, t in zip(("date", "type", "number", "qty", "balance"),
                            ("تاریخ", "نوع حرکت", "شماره فاکتور", "تعداد", "مانده انباری تجمعی")):
                ltree.heading(c, text=t)
                ltree.column(c, width=120, anchor="center")
            ltree.pack(fill="both", expand=True, padx=10, pady=10)
            for row in data["ledger"]:
                ltree.insert("", "end", values=(row["date"], row["type"], row["number"], row["qty"], row["running_balance"]))

        def export_items_excel():
            headers = ["کد/بارکد", "نام", "برند", "واحد", "قیمت خرید", "قیمت فروش", "موجودی", "حداقل موجودی"]
            rows = []
            for it in getattr(self, "all_items", []):
                rows.append([it.get("code") or "", it["name"], it.get("brand") or "", it["unit"],
                             it["purchase_price"], it["sale_price"], it["stock_qty"], it["min_stock"]])
            self.export_to_excel(headers, rows, "کالاها")

        btn_row = ttk.Frame(parent)
        btn_row.pack(pady=5)
        ttk.Button(btn_row, text="بروزرسانی لیست", command=self.refresh_items).pack(side="right", padx=5)
        ttk.Button(btn_row, text="مشاهده گردش انبار کالای انتخاب‌شده", command=show_stock_ledger).pack(side="right", padx=5)
        ttk.Button(btn_row, text="خروجی اکسل از لیست کالاها", command=export_items_excel).pack(side="right", padx=5)
        self.refresh_items()
        self.tab_refresh_callbacks[parent] = self.refresh_items

    def refresh_items(self):
        for r in self.items_tree.get_children():
            self.items_tree.delete(r)
        items = self.api("GET", "/items")
        if not items:
            return
        for it in items:
            self.items_tree.insert("", "end", values=(
                it["id"], it["code"] or "", it["name"], it.get("brand") or "", it["unit"],
                it["purchase_price"], it["sale_price"], it["stock_qty"], it["min_stock"]
            ))
        # برای فرم فاکتورها هم لیست کالاها را نگه می‌داریم
        self.all_items = items

    def add_item(self):
        e = self.item_entries
        cat_name = self.item_category_combo.get()
        category_id = None
        if cat_name:
            for c in getattr(self, "categories_list", []):
                if c["name"] == cat_name:
                    category_id = c["id"]
                    break
        try:
            payload = {
                "code": e["کد/بارکد"].get(),
                "name": e["نام کالا"].get(),
                "brand": e["برند"].get(),
                "unit": e["واحد"].get() or "عدد",
                "purchase_price": float(e["قیمت خرید"].get() or 0),
                "sale_price": float(e["قیمت فروش"].get() or 0),
                "stock_qty": float(e["موجودی"].get() or 0),
                "min_stock": float(e["حداقل موجودی"].get() or 0),
                "category_id": category_id,
            }
        except ValueError:
            messagebox.showerror("خطا", "مقادیر عددی را درست وارد کنید")
            return
        if not payload["name"]:
            messagebox.showerror("خطا", "نام کالا الزامی است")
            return
        res = self.api("POST", "/items", json=payload)
        if res and res.get("ok"):
            messagebox.showinfo("موفق", "کالا اضافه شد")
            self.refresh_items()

    # ---------------- فاکتور فروش/خرید ----------------
    def build_invoice_tab(self, parent, invoice_type):
        self.all_items = getattr(self, "all_items", [])
        cart_key = f"cart_{invoice_type}"
        setattr(self, cart_key, [])

        desc = ("وقتی مشتری از مغازه‌ات جنس می‌خرد اینجا ثبت کن — مثلاً «۲ عدد لامپ نقدی فروختی». موجودی خودکار کم می‌شود."
                if invoice_type == "sale" else
                "وقتی خودت از تامین‌کننده جنس می‌خری تا انبار پر بشه اینجا ثبت کن — مثلاً «۵۰ عدد لامپ خریدی». "
                "کالا باید قبلاً در تب «کالاها» تعریف شده باشد تا این‌جا در لیست ظاهر شود.")
        ttk.Label(parent, text=desc, foreground="gray", wraplength=850, padding=(10, 8, 10, 0)).pack(anchor="w")

        top = ttk.Frame(parent, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="طرف حساب:").grid(row=0, column=3, padx=5)
        party_combo = ttk.Combobox(top, width=20, state="readonly")
        party_combo.grid(row=0, column=2, padx=5)
        setattr(self, f"party_combo_{invoice_type}", party_combo)

        ttk.Label(top, text="نوع پرداخت:").grid(row=0, column=1, padx=5)
        pay_combo = ttk.Combobox(top, width=10, state="readonly", values=["cash", "credit", "check"])
        pay_combo.set("cash")
        pay_combo.grid(row=0, column=0, padx=5)
        setattr(self, f"pay_combo_{invoice_type}", pay_combo)

        ttk.Label(top, text="کالا:").grid(row=1, column=3, padx=5, pady=8)
        item_combo = ttk.Combobox(top, width=25, state="readonly")
        item_combo.grid(row=1, column=2, padx=5)
        setattr(self, f"item_combo_{invoice_type}", item_combo)

        ttk.Label(top, text="تعداد/مقدار:").grid(row=1, column=1, padx=5)
        qty_entry = ttk.Entry(top, width=10)
        qty_entry.grid(row=1, column=0, padx=5)

        ttk.Label(top, text="قیمت واحد:").grid(row=2, column=1, padx=5)
        price_entry = ttk.Entry(top, width=10)
        price_entry.grid(row=2, column=0, padx=5)

        ttk.Label(top, text="مبلغ تخفیف کل فاکتور:").grid(row=2, column=3, padx=5)
        discount_entry = ttk.Entry(top, width=12)
        discount_entry.insert(0, "0")
        discount_entry.grid(row=2, column=4, padx=5)

        is_return_var = tk.BooleanVar(value=False)
        return_label = "این یک فاکتور مرجوعی فروش است (مشتری کالا پس آورده)" if invoice_type == "sale" \
            else "این یک فاکتور مرجوعی خرید است (کالا به تامین‌کننده پس داده شده)"
        ttk.Checkbutton(top, text=return_label, variable=is_return_var).grid(row=3, column=0, columnspan=3, sticky="w", pady=5)

        def refresh_combos():
            parties = self.api("GET", "/parties") or []
            party_combo["values"] = [f'{p["id"]} - {p["name"]}' for p in parties]
            items = self.api("GET", "/items") or []
            self.all_items = items
            item_combo["values"] = [f'{it["id"]} - {it["name"]}' for it in items]

        refresh_combos()
        self.tab_refresh_callbacks[parent] = refresh_combos
        ttk.Button(top, text="بروزرسانی لیست‌ها", command=refresh_combos).grid(row=2, column=2)

        cart_tree = ttk.Treeview(parent, columns=("name", "qty", "price", "total"), show="headings", height=8)
        for c, t in zip(("name", "qty", "price", "total"), ("کالا", "تعداد", "قیمت واحد", "جمع")):
            cart_tree.heading(c, text=t)
            cart_tree.column(c, width=150, anchor="center")
        cart_tree.pack(fill="x", padx=10, pady=10)
        setattr(self, f"cart_tree_{invoice_type}", cart_tree)

        def add_to_cart():
            if not item_combo.get():
                messagebox.showerror("خطا", "یک کالا انتخاب کنید")
                return
            try:
                qty = float(qty_entry.get())
                price = float(price_entry.get())
            except ValueError:
                messagebox.showerror("خطا", "تعداد و قیمت را درست وارد کنید")
                return
            item_id = int(item_combo.get().split(" - ")[0])
            item_name = item_combo.get().split(" - ", 1)[1]
            cart = getattr(self, cart_key)
            cart.append({"item_id": item_id, "qty": qty, "unit_price": price})
            cart_tree.insert("", "end", values=(item_name, qty, price, qty * price))

        ttk.Button(top, text="افزودن به فاکتور", command=add_to_cart).grid(row=1, column=4, padx=10)

        def submit_invoice():
            cart = getattr(self, cart_key)
            if not cart:
                messagebox.showerror("خطا", "فاکتور خالی است")
                return
            party_val = party_combo.get()
            party_id = int(party_val.split(" - ")[0]) if party_val else None
            payment_type = pay_combo.get() or "cash"
            try:
                discount = float(discount_entry.get() or 0)
            except ValueError:
                messagebox.showerror("خطا", "مبلغ تخفیف را به‌صورت عدد وارد کنید")
                return

            effective_type = invoice_type
            if is_return_var.get():
                effective_type = "sale_return" if invoice_type == "sale" else "purchase_return"

            payload = {
                "invoice_type": effective_type,
                "party_id": party_id,
                "payment_type": payment_type,
                "discount": discount,
                "username": self.user["username"] if self.user else None,
                "items": cart
            }
            res = self.api("POST", "/invoices", json=payload)
            if res and res.get("ok"):
                invoice_id = res["invoice_id"]
                invoice_number = res.get("invoice_number", str(invoice_id))
                total = res["total"]
                messagebox.showinfo("موفق", f'فاکتور {invoice_number} ثبت شد. جمع کل: {total:,.0f} تومان')

                # اگر پرداخت با چک بود، اطلاعات چک را می‌گیریم و ثبت می‌کنیم
                if payment_type == "check":
                    due_date = simpledialog.askstring(
                        "تاریخ سررسید چک", "تاریخ سررسید چک را وارد کنید (مثلاً 1404-06-01):")
                    if due_date:
                        direction = "received" if invoice_type == "sale" else "issued"
                        self.api("POST", "/checks", json={
                            "party_id": party_id, "invoice_id": invoice_id,
                            "amount": total, "due_date": due_date,
                            "direction": direction, "description": f"چک فاکتور شماره {invoice_number}"
                        })

                # پیشنهاد چاپ فاکتور
                if messagebox.askyesno("چاپ فاکتور", "آیا مایلید فاکتور را به‌صورت PDF چاپ/ذخیره کنید؟"):
                    print_items = []
                    for row_id in cart_tree.get_children():
                        vals = cart_tree.item(row_id)["values"]
                        print_items.append({"item_name": vals[0], "qty": vals[1],
                                             "unit_price": vals[2], "total": vals[3]})
                    self.print_invoice_pdf(
                        {"id": invoice_number, "invoice_type": effective_type, "date": "", "total": total,
                         "paid": total if payment_type == "cash" else 0},
                        print_items, party_val.split(" - ", 1)[1] if party_val else None
                    )

                setattr(self, cart_key, [])
                for r in cart_tree.get_children():
                    cart_tree.delete(r)
                discount_entry.delete(0, tk.END)
                discount_entry.insert(0, "0")
                is_return_var.set(False)
                self.refresh_items()

        ttk.Button(parent, text="ثبت نهایی فاکتور", command=submit_invoice).pack(pady=10)

    # ---------------- مشتریان و تامین‌کنندگان ----------------
    def build_parties_tab(self, parent):
        ttk.Label(parent, text="افرادی که از تو نسیه می‌خرن (مشتری) یا تو ازشون نسیه می‌خری (تامین‌کننده) — مثلاً «آقای رضایی ۸۰,۰۰۰ تومان بدهکاره».",
                  foreground="gray", padding=(10, 8, 10, 0)).pack(anchor="w")
        top = ttk.Frame(parent, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="نام:").grid(row=0, column=3)
        name_e = ttk.Entry(top, width=20)
        name_e.grid(row=0, column=2, padx=5)

        ttk.Label(top, text="تلفن:").grid(row=0, column=1)
        phone_e = ttk.Entry(top, width=15)
        phone_e.grid(row=0, column=0, padx=5)

        ttk.Label(top, text="نوع:").grid(row=1, column=1)
        type_combo = ttk.Combobox(top, width=15, state="readonly", values=["customer", "supplier"])
        type_combo.set("customer")
        type_combo.grid(row=1, column=0, padx=5)

        tree = ttk.Treeview(parent, columns=("id", "name", "phone", "type", "balance"), show="headings")
        for c, t in zip(("id", "name", "phone", "type", "balance"),
                        ("ردیف", "نام", "تلفن", "نوع", "مانده حساب (مثبت=بدهکار)")):
            tree.heading(c, text=t)
            tree.column(c, width=140, anchor="center")
        tree.pack(fill="both", expand=True, padx=10, pady=10)

        def refresh():
            for r in tree.get_children():
                tree.delete(r)
            parties = self.api("GET", "/parties") or []
            for p in parties:
                tree.insert("", "end", values=(p["id"], p["name"], p["phone"] or "", p["type"], p["balance"]))

        def add_party():
            if not name_e.get():
                messagebox.showerror("خطا", "نام الزامی است")
                return
            res = self.api("POST", "/parties", json={
                "name": name_e.get(), "phone": phone_e.get(), "type": type_combo.get()
            })
            if res and res.get("ok"):
                refresh()

        def settle_payment():
            sel = tree.selection()
            if not sel:
                messagebox.showerror("خطا", "یک طرف‌حساب را از لیست انتخاب کنید")
                return
            vals = tree.item(sel[0])["values"]
            party_id, party_name = vals[0], vals[1]
            amount_str = simpledialog.askstring("تسویه حساب", f'مبلغ دریافتی/پرداختی برای «{party_name}» را وارد کنید:')
            if not amount_str:
                return
            try:
                amount = float(amount_str)
            except ValueError:
                messagebox.showerror("خطا", "مبلغ را به‌صورت عدد وارد کنید")
                return
            res = self.api("POST", f"/parties/{party_id}/payment", json={"amount": amount})
            if res and res.get("ok"):
                messagebox.showinfo("موفق", f'ثبت شد. مانده جدید حساب: {res["new_balance"]:,.0f} تومان')
                refresh()

        def show_ledger():
            sel = tree.selection()
            if not sel:
                messagebox.showerror("خطا", "یک طرف‌حساب را از لیست انتخاب کنید")
                return
            party_id = tree.item(sel[0])["values"][0]
            data = self.api("GET", f"/parties/{party_id}/ledger")
            if not data:
                return
            win = tk.Toplevel(self)
            win.title(f'ریز حساب: {data["party"]["name"]}')
            win.geometry("600x400")
            ttk.Label(win, text=f'مانده فعلی حساب: {data["party"]["balance"]:,.0f} تومان',
                      font=("Tahoma", 12, "bold")).pack(pady=10)
            ltree = ttk.Treeview(win, columns=("date", "type", "total", "paid"), show="headings")
            for c, t in zip(("date", "type", "total", "paid"), ("تاریخ", "نوع فاکتور", "جمع کل", "پرداخت‌شده")):
                ltree.heading(c, text=t)
                ltree.column(c, width=130, anchor="center")
            ltree.pack(fill="both", expand=True, padx=10, pady=10)
            for inv in data["invoices"]:
                ltree.insert("", "end", values=(
                    inv["date"], "فروش" if inv["invoice_type"] == "sale" else "خرید",
                    inv["total"], inv["paid"]
                ))

        ttk.Button(top, text="افزودن", command=add_party).grid(row=1, column=2, padx=5)
        button_row = ttk.Frame(parent)
        button_row.pack(pady=5)
        ttk.Button(button_row, text="بروزرسانی", command=refresh).pack(side="right", padx=5)
        ttk.Button(button_row, text="ثبت تسویه حساب (دریافت/پرداخت)", command=settle_payment).pack(side="right", padx=5)
        ttk.Button(button_row, text="مشاهده ریز حساب", command=show_ledger).pack(side="right", padx=5)
        refresh()
        self.tab_refresh_callbacks[parent] = refresh

    # ---------------- صندوق ----------------
    def build_cash_tab(self, parent):
        top = ttk.Frame(parent, padding=10)
        top.pack(fill="x")

        self.cash_balance_label = ttk.Label(top, text="موجودی صندوق: -", font=("Tahoma", 13, "bold"))
        self.cash_balance_label.grid(row=0, column=0, columnspan=4, pady=10)

        ttk.Label(top, text="مبلغ:").grid(row=1, column=3)
        amount_e = ttk.Entry(top, width=15)
        amount_e.grid(row=1, column=2, padx=5)

        ttk.Label(top, text="نوع:").grid(row=1, column=1)
        type_combo = ttk.Combobox(top, width=10, state="readonly", values=["in", "out"])
        type_combo.set("in")
        type_combo.grid(row=1, column=0, padx=5)

        ttk.Label(top, text="توضیحات:").grid(row=2, column=1)
        desc_e = ttk.Entry(top, width=30)
        desc_e.grid(row=2, column=0, columnspan=2, padx=5)

        tree = ttk.Treeview(parent, columns=("date", "type", "amount", "desc"), show="headings")
        for c, t in zip(("date", "type", "amount", "desc"), ("تاریخ", "نوع", "مبلغ", "توضیحات")):
            tree.heading(c, text=t)
            tree.column(c, width=150, anchor="center")
        tree.pack(fill="both", expand=True, padx=10, pady=10)

        def refresh():
            data = self.api("GET", "/cash")
            if not data:
                return
            self.cash_balance_label.config(text=f'موجودی صندوق: {data["balance"]:,.0f} تومان')
            for r in tree.get_children():
                tree.delete(r)
            for tx in data["transactions"]:
                tree.insert("", "end", values=(
                    tx["date"], "دریافت" if tx["tx_type"] == "in" else "پرداخت",
                    tx["amount"], tx["description"] or ""
                ))

        def add_tx():
            try:
                amount = float(amount_e.get())
            except ValueError:
                messagebox.showerror("خطا", "مبلغ را درست وارد کنید")
                return
            res = self.api("POST", "/cash", json={
                "tx_type": type_combo.get(), "amount": amount, "description": desc_e.get(),
                "username": self.user["username"] if self.user else None
            })
            if res and res.get("ok"):
                refresh()

        ttk.Button(top, text="ثبت تراکنش", command=add_tx).grid(row=2, column=2, padx=5)
        ttk.Button(parent, text="بروزرسانی", command=refresh).pack(pady=5)
        refresh()
        self.tab_refresh_callbacks[parent] = refresh

    # ---------------- تاریخچه و چاپ فاکتورها ----------------
    def build_history_tab(self, parent):
        top = ttk.Frame(parent, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="نوع فاکتور:").grid(row=0, column=1, padx=5)
        type_combo = ttk.Combobox(top, width=12, state="readonly", values=["همه", "sale", "purchase"])
        type_combo.set("همه")
        type_combo.grid(row=0, column=0, padx=5)

        tree = ttk.Treeview(parent, columns=("id", "type", "date", "party", "total", "paid", "description"), show="headings")
        headers = {"id": "شماره", "type": "نوع", "date": "تاریخ", "party": "طرف حساب",
                   "total": "جمع کل", "paid": "پرداخت‌شده", "description": "توضیحات"}
        for c in headers:
            tree.heading(c, text=headers[c])
            tree.column(c, width=120, anchor="center")
        tree.pack(fill="both", expand=True, padx=10, pady=10)

        type_label_fa = {"sale": "فروش", "purchase": "خرید", "sale_return": "مرجوعی فروش", "purchase_return": "مرجوعی خرید"}
        self._history_invoices = {}

        def refresh():
            for r in tree.get_children():
                tree.delete(r)
            t = None if type_combo.get() == "همه" else type_combo.get()
            path = "/invoices" + (f"?type={t}" if t else "")
            invoices = self.api("GET", path) or []
            self._history_invoices = {inv["id"]: inv for inv in invoices}
            for inv in invoices:
                tree.insert("", "end", values=(
                    inv["id"], type_label_fa.get(inv["invoice_type"], inv["invoice_type"]),
                    inv["date"], inv.get("party_name") or "-", inv["total"], inv["paid"], inv.get("description") or "-"
                ))

        def edit_invoice_full():
            sel = tree.selection()
            if not sel:
                messagebox.showerror("خطا", "یک فاکتور از لیست انتخاب کنید")
                return
            invoice_id = tree.item(sel[0])["values"][0]
            inv = self._history_invoices.get(invoice_id)
            if not inv:
                messagebox.showerror("خطا", "اطلاعات فاکتور پیدا نشد؛ لیست را بروزرسانی کنید")
                return

            inv_items = self.api("GET", f"/invoices/{invoice_id}/items") or []
            all_items = self.api("GET", "/items") or []
            all_parties = self.api("GET", "/parties") or []
            item_by_id = {it["id"]: it for it in all_items}
            wanted_party_type = "customer" if inv["invoice_type"] in ("sale", "sale_return") else "supplier"
            parties = [p for p in all_parties if p["type"] == wanted_party_type]

            cart = [{"item_id": it["item_id"], "item_name": it["item_name"],
                     "qty": it["qty"], "unit_price": it["unit_price"]} for it in inv_items]

            win = tk.Toplevel(self)
            win.title(f'ویرایش فاکتور {inv.get("number") or invoice_id}')
            win.geometry("560x560")

            ttk.Label(win, text=f'نوع فاکتور ({type_label_fa.get(inv["invoice_type"], inv["invoice_type"])}) قابل تغییر نیست. '
                                 "موجودی انبار و حساب طرف‌حساب بر اساس مقادیر جدید دوباره محاسبه می‌شود.",
                      wraplength=520, foreground="gray").pack(padx=15, pady=(15, 10))

            top_frame = ttk.Frame(win, padding=(15, 0))
            top_frame.pack(fill="x")

            ttk.Label(top_frame, text="طرف‌حساب:").grid(row=0, column=0, sticky="e", padx=5, pady=3)
            party_names = ["— بدون طرف‌حساب —"] + [p["name"] for p in parties]
            party_combo = ttk.Combobox(top_frame, state="readonly", values=party_names, width=25)
            current_party = next((p["name"] for p in parties if p["id"] == inv.get("party_id")), None)
            party_combo.set(current_party or party_names[0])
            party_combo.grid(row=0, column=1, padx=5, pady=3)

            ttk.Label(top_frame, text="نوع پرداخت:").grid(row=0, column=2, sticky="e", padx=5, pady=3)
            pay_fa = {"cash": "نقدی", "credit": "نسیه", "check": "چک"}
            pay_combo = ttk.Combobox(top_frame, state="readonly", values=["نقدی", "نسیه", "چک"], width=10)
            pay_combo.set(pay_fa.get(inv["payment_type"], "نقدی"))
            pay_combo.grid(row=0, column=3, padx=5, pady=3)

            ttk.Label(top_frame, text="تخفیف کل (تومان):").grid(row=1, column=0, sticky="e", padx=5, pady=3)
            discount_entry = ttk.Entry(top_frame, width=15)
            discount_entry.insert(0, str(inv.get("discount", 0) or 0))
            discount_entry.grid(row=1, column=1, padx=5, pady=3)

            ttk.Label(top_frame, text="توضیحات:").grid(row=1, column=2, sticky="e", padx=5, pady=3)
            desc_entry = ttk.Entry(top_frame, width=25)
            desc_entry.insert(0, inv.get("description") or "")
            desc_entry.grid(row=1, column=3, padx=5, pady=3)

            ttk.Label(win, text="اقلام فاکتور:", font=("Tahoma", 10, "bold")).pack(padx=15, pady=(10, 0), anchor="e")

            cart_tree = ttk.Treeview(win, columns=("item", "qty", "price", "total"), show="headings", height=6)
            for c, t in zip(("item", "qty", "price", "total"), ("کالا", "تعداد", "قیمت واحد", "جمع")):
                cart_tree.heading(c, text=t)
                cart_tree.column(c, width=110, anchor="center")
            cart_tree.pack(fill="both", expand=True, padx=15, pady=5)

            def refresh_cart_tree():
                for r in cart_tree.get_children():
                    cart_tree.delete(r)
                for row in cart:
                    cart_tree.insert("", "end", values=(row["item_name"], row["qty"], row["unit_price"],
                                                          row["qty"] * row["unit_price"]))

            refresh_cart_tree()

            def remove_selected_row():
                sel2 = cart_tree.selection()
                if not sel2:
                    return
                idx = cart_tree.index(sel2[0])
                cart.pop(idx)
                refresh_cart_tree()

            ttk.Button(win, text="حذف ردیف انتخاب‌شده", command=remove_selected_row).pack(padx=15, pady=2, anchor="e")

            add_frame = ttk.Frame(win, padding=(15, 5))
            add_frame.pack(fill="x")
            price_field = "sale_price" if inv["invoice_type"] in ("sale", "sale_return") else "purchase_price"
            item_names = [it["name"] for it in all_items]
            ttk.Label(add_frame, text="کالا:").grid(row=0, column=0, sticky="e", padx=3)
            item_combo = ttk.Combobox(add_frame, values=item_names, width=20)
            item_combo.grid(row=0, column=1, padx=3)
            ttk.Label(add_frame, text="تعداد:").grid(row=0, column=2, sticky="e", padx=3)
            qty_entry = ttk.Entry(add_frame, width=8)
            qty_entry.grid(row=0, column=3, padx=3)
            ttk.Label(add_frame, text="قیمت واحد:").grid(row=0, column=4, sticky="e", padx=3)
            price_entry = ttk.Entry(add_frame, width=10)
            price_entry.grid(row=0, column=5, padx=3)

            def add_line():
                name = item_combo.get()
                match = next((it for it in all_items if it["name"] == name), None)
                if not match:
                    messagebox.showerror("خطا", "کالای انتخاب‌شده معتبر نیست")
                    return
                try:
                    qty = float(qty_entry.get())
                    price = float(price_entry.get())
                except ValueError:
                    messagebox.showerror("خطا", "تعداد و قیمت باید عدد باشند")
                    return
                if not qty or not price:
                    messagebox.showerror("خطا", "تعداد و قیمت را وارد کنید")
                    return
                cart.append({"item_id": match["id"], "item_name": match["name"], "qty": qty, "unit_price": price})
                qty_entry.delete(0, tk.END)
                price_entry.delete(0, tk.END)
                refresh_cart_tree()

            def on_item_selected(event):
                name = item_combo.get()
                match = next((it for it in all_items if it["name"] == name), None)
                if match and match.get(price_field):
                    price_entry.delete(0, tk.END)
                    price_entry.insert(0, str(match[price_field]))

            item_combo.bind("<<ComboboxSelected>>", on_item_selected)
            ttk.Button(add_frame, text="+ افزودن قلم", command=add_line).grid(row=0, column=6, padx=8)

            def save():
                if not cart:
                    messagebox.showerror("خطا", "فاکتور باید حداقل یک کالا داشته باشد")
                    return
                if not messagebox.askyesno("تأیید ویرایش",
                        "آیا مطمئنی می‌خوای این فاکتور رو با این تغییرات ذخیره کنی؟\n"
                        "موجودی انبار و حساب طرف‌حساب بر این اساس دوباره محاسبه می‌شود."):
                    return
                selected_party = next((p for p in parties if p["name"] == party_combo.get()), None)
                pay_type = {v: k for k, v in pay_fa.items()}.get(pay_combo.get(), "cash")
                try:
                    discount = float(discount_entry.get() or 0)
                except ValueError:
                    discount = 0
                payload = {
                    "party_id": selected_party["id"] if selected_party else None,
                    "payment_type": pay_type,
                    "discount": discount,
                    "description": desc_entry.get(),
                    "username": self.user["username"],
                    "items": [{"item_id": r["item_id"], "qty": r["qty"], "unit_price": r["unit_price"]} for r in cart],
                }
                res = self.api("PUT", f"/invoices/{invoice_id}", json=payload)
                if res and res.get("ok"):
                    messagebox.showinfo("موفق", "فاکتور با موفقیت ویرایش شد")
                    win.destroy()
                    refresh()
                else:
                    messagebox.showerror("خطا", (res or {}).get("message", "ذخیره نشد"))

            ttk.Button(win, text="ذخیره تغییرات", command=save).pack(pady=15)

        def print_selected():
            sel = tree.selection()
            if not sel:
                messagebox.showerror("خطا", "یک فاکتور از لیست انتخاب کنید")
                return
            vals = tree.item(sel[0])["values"]
            invoice_id = vals[0]
            invoice_type = "sale" if vals[1] == "فروش" else "purchase"
            items = self.api("GET", f"/invoices/{invoice_id}/items") or []
            invoice = {"id": invoice_id, "invoice_type": invoice_type, "date": vals[2],
                       "total": vals[4], "paid": vals[5]}
            party_name = vals[3] if vals[3] != "-" else None
            self.print_invoice_pdf(invoice, items, party_name)

        ttk.Button(top, text="بروزرسانی لیست", command=refresh).grid(row=0, column=2, padx=10)

        def export_excel():
            headers = ["شماره", "نوع", "تاریخ", "طرف حساب", "جمع کل", "پرداخت‌شده"]
            rows = [tree.item(r)["values"] for r in tree.get_children()]
            self.export_to_excel(headers, rows, "فاکتورها")

        btn_row = ttk.Frame(parent)
        btn_row.pack(pady=5)
        ttk.Button(btn_row, text="چاپ / ذخیره PDF فاکتور انتخاب‌شده", command=print_selected).pack(side="right", padx=5)
        if self.user and self.user.get("role") == "admin":
            ttk.Button(btn_row, text="ویرایش فاکتور انتخاب‌شده", command=edit_invoice_full).pack(side="right", padx=5)
        ttk.Button(btn_row, text="خروجی اکسل از این لیست", command=export_excel).pack(side="right", padx=5)
        refresh()
        self.tab_refresh_callbacks[parent] = refresh

    # ---------------- چک‌ها ----------------
    def build_checks_tab(self, parent):
        top = ttk.Frame(parent, padding=10)
        top.pack(fill="x")

        tree = ttk.Treeview(parent, columns=("id", "party", "amount", "due", "direction", "status"), show="headings")
        headers = {"id": "شماره", "party": "طرف حساب", "amount": "مبلغ",
                   "due": "سررسید", "direction": "نوع", "status": "وضعیت"}
        for c in headers:
            tree.heading(c, text=headers[c])
            tree.column(c, width=130, anchor="center")
        tree.pack(fill="both", expand=True, padx=10, pady=10)

        status_fa = {"pending": "در انتظار", "cashed": "وصول‌شده", "bounced": "برگشتی"}
        direction_fa = {"received": "دریافتی از مشتری", "issued": "صادرشده به تامین‌کننده"}

        def refresh():
            for r in tree.get_children():
                tree.delete(r)
            checks = self.api("GET", "/checks") or []
            for ch in checks:
                tree.insert("", "end", values=(
                    ch["id"], ch.get("party_name") or "-", ch["amount"], ch["due_date"],
                    direction_fa.get(ch["direction"], ch["direction"]),
                    status_fa.get(ch["status"], ch["status"])
                ))

        def set_status(new_status):
            sel = tree.selection()
            if not sel:
                messagebox.showerror("خطا", "یک چک را از لیست انتخاب کنید")
                return
            check_id = tree.item(sel[0])["values"][0]
            res = self.api("PUT", f"/checks/{check_id}", json={"status": new_status})
            if res and res.get("ok"):
                refresh()

        ttk.Button(top, text="بروزرسانی لیست", command=refresh).grid(row=0, column=3, padx=5)
        ttk.Button(top, text="علامت‌گذاری به‌عنوان وصول‌شده", command=lambda: set_status("cashed")).grid(row=0, column=2, padx=5)
        ttk.Button(top, text="علامت‌گذاری به‌عنوان برگشتی", command=lambda: set_status("bounced")).grid(row=0, column=1, padx=5)
        ttk.Label(parent, text="توجه: چک‌های ثبت‌شده هنگام صدور فاکتور با پرداخت «check» به‌صورت خودکار اینجا اضافه می‌شوند.",
                  foreground="gray").pack(pady=5)
        refresh()
        self.tab_refresh_callbacks[parent] = refresh

    # ---------------- پشتیبان‌گیری ----------------
    def build_backup_tab(self, parent):
        top = ttk.Frame(parent, padding=20)
        top.pack(fill="x")

        ttk.Label(top, text="پشتیبان‌گیری خودکار روزانه از دیتابیس روی سرور فعال است.",
                  font=("Tahoma", 11, "bold")).pack(pady=5)
        ttk.Label(top, text="در صورت نیاز می‌توانید همین حالا هم یک نسخه پشتیبان بگیرید:").pack(pady=5)

        def backup_now():
            res = self.api("POST", "/backup/now")
            if res and res.get("ok"):
                messagebox.showinfo("موفق", f'نسخه پشتیبان ساخته شد:\n{res.get("path")}')
                refresh_list()

        ttk.Button(top, text="تهیه نسخه پشتیبان الان", command=backup_now).pack(pady=10)

        ttk.Label(top, text="گزارش شبانه (ایمیل خلاصه فروش + آپلود بکاپ به VPS) هر شب ساعت ۲۳ خودکار اجرا می‌شود.",
                  font=("Tahoma", 10)).pack(pady=5)

        def run_nightly_now():
            res = self.api("POST", "/nightly/run-now")
            if res:
                email_msg = res.get("email", {}).get("message", "")
                vps_msg = res.get("vps", {}).get("message", "")
                messagebox.showinfo("نتیجه گزارش شبانه", f"ایمیل: {email_msg}\n\nVPS: {vps_msg}")

        ttk.Button(top, text="ارسال گزارش شبانه همین الان (تست)", command=run_nightly_now).pack(pady=5)

        list_frame = ttk.Frame(parent, padding=10)
        list_frame.pack(fill="both", expand=True)
        listbox = tk.Listbox(list_frame)
        listbox.pack(fill="both", expand=True)

        def refresh_list():
            listbox.delete(0, tk.END)
            files = self.api("GET", "/backup/list") or []
            for f in files:
                listbox.insert(tk.END, f)

        def restore_selected():
            sel = listbox.curselection()
            if not sel:
                messagebox.showerror("خطا", "یک نسخه پشتیبان را از لیست انتخاب کنید")
                return
            filename = listbox.get(sel[0])
            if not messagebox.askyesno(
                "تأیید بازیابی",
                f'آیا مطمئنید می‌خواهید دیتابیس فعلی با نسخه «{filename}» جایگزین شود؟\n'
                'یک نسخه احتیاطی از وضعیت فعلی هم قبل از این کار گرفته می‌شود، ولی این عمل قابل بازگشت آسان نیست.'
            ):
                return
            res = self.api("POST", "/backup/restore", json={
                "filename": filename, "username": self.user["username"] if self.user else None
            })
            if res:
                if res.get("ok"):
                    messagebox.showinfo("موفق", res.get("message", "بازیابی انجام شد"))
                    refresh_list()
                else:
                    messagebox.showerror("خطا", res.get("message", "خطا در بازیابی"))

        ttk.Button(parent, text="بروزرسانی لیست نسخه‌های پشتیبان", command=refresh_list).pack(pady=5)
        ttk.Button(parent, text="بازیابی از نسخه انتخاب‌شده", command=restore_selected).pack(pady=5)
        refresh_list()
        self.tab_refresh_callbacks[parent] = refresh_list

    # ---------------- گزارش ماهانه (با نمودار) ----------------
    def build_monthly_tab(self, parent):
        top = ttk.Frame(parent, padding=10)
        top.pack(fill="x")

        fig = Figure(figsize=(7, 3.2), dpi=100)
        ax = fig.add_subplot(111)
        canvas = FigureCanvasTkAgg(fig, master=parent)
        canvas.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=10)

        tree = ttk.Treeview(parent, columns=("month", "sales", "purchases", "profit"), show="headings", height=6)
        headers = {"month": "ماه", "sales": "فروش", "purchases": "خرید", "profit": "سود/زیان"}
        for c in headers:
            tree.heading(c, text=headers[c])
            tree.column(c, width=140, anchor="center")
        tree.pack(fill="x", padx=10, pady=5)

        def refresh():
            data = self.api("GET", "/reports/monthly") or []
            for r in tree.get_children():
                tree.delete(r)
            for row in data:
                tree.insert("", "end", values=(row["month"], f'{row["sales"]:,.0f}',
                                                f'{row["purchases"]:,.0f}', f'{row["profit"]:,.0f}'))

            ax.clear()
            if data:
                months = [d["month"] for d in data]
                profits = [d["profit"] for d in data]
                colors = ["#1D9E75" if p >= 0 else "#D85A30" for p in profits]
                ax.bar(months, profits, color=colors)
                ax.set_title("سود/زیان ماهانه (تومان)")
                ax.tick_params(axis="x", rotation=45)
                fig.tight_layout()
            canvas.draw()

        ttk.Button(top, text="بروزرسانی گزارش", command=refresh).pack(pady=5)
        refresh()
        self.tab_refresh_callbacks[parent] = refresh

    # ---------------- بدهکاران، بستانکاران، پرفروش‌ترین کالاها ----------------
    def build_extra_reports_tab(self, parent):
        # --- مشتریان بدهکار ---
        ttk.Label(parent, text="مشتریان بدهکار به ما", font=("Tahoma", 11, "bold")).pack(pady=(10, 2))
        debtors_tree = ttk.Treeview(parent, columns=("name", "phone", "balance"), show="headings", height=5)
        for c, t in zip(("name", "phone", "balance"), ("نام مشتری", "تلفن", "مبلغ بدهی")):
            debtors_tree.heading(c, text=t)
            debtors_tree.column(c, width=150, anchor="center")
        debtors_tree.pack(fill="x", padx=10, pady=5)

        # --- تامین‌کنندگان بستانکار ---
        ttk.Label(parent, text="تامین‌کنندگانی که به آن‌ها بدهکاریم", font=("Tahoma", 11, "bold")).pack(pady=(10, 2))
        creditors_tree = ttk.Treeview(parent, columns=("name", "phone", "owed"), show="headings", height=5)
        for c, t in zip(("name", "phone", "owed"), ("نام تامین‌کننده", "تلفن", "مبلغ بدهی ما")):
            creditors_tree.heading(c, text=t)
            creditors_tree.column(c, width=150, anchor="center")
        creditors_tree.pack(fill="x", padx=10, pady=5)

        # --- پرفروش‌ترین کالاها ---
        ttk.Label(parent, text="پرفروش‌ترین کالاها (۳۰ روز اخیر)", font=("Tahoma", 11, "bold")).pack(pady=(10, 2))
        top_tree = ttk.Treeview(parent, columns=("name", "brand", "qty", "amount"), show="headings", height=6)
        for c, t in zip(("name", "brand", "qty", "amount"), ("نام کالا", "برند", "تعداد فروخته‌شده", "جمع فروش")):
            top_tree.heading(c, text=t)
            top_tree.column(c, width=140, anchor="center")
        top_tree.pack(fill="both", expand=True, padx=10, pady=5)

        def refresh():
            for tree, endpoint, cols in [
                (debtors_tree, "/reports/debtors", ("name", "phone", "balance")),
                (creditors_tree, "/reports/creditors", ("name", "phone", "owed_amount")),
            ]:
                for r in tree.get_children():
                    tree.delete(r)
                data = self.api("GET", endpoint) or []
                for row in data:
                    tree.insert("", "end", values=tuple(row.get(c, "") for c in cols))

            for r in top_tree.get_children():
                top_tree.delete(r)
            top_data = self.api("GET", "/reports/top-items") or []
            for row in top_data:
                top_tree.insert("", "end", values=(row["name"], row.get("brand") or "", row["total_qty"], row["total_amount"]))

        def export_all_excel():
            debtors = self.api("GET", "/reports/debtors") or []
            creditors = self.api("GET", "/reports/creditors") or []
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "بدهکاران"
            ws1.append(["نام", "تلفن", "مبلغ بدهی"])
            for p in debtors:
                ws1.append([p["name"], p.get("phone") or "", p["balance"]])
            ws2 = wb.create_sheet("بستانکاران")
            ws2.append(["نام", "تلفن", "مبلغ بدهی ما"])
            for p in creditors:
                ws2.append([p["name"], p.get("phone") or "", p["owed_amount"]])
            out_path = os.path.join(tempfile.gettempdir(), "بدهکاران_بستانکاران.xlsx")
            try:
                wb.save(out_path)
                if sys.platform.startswith("win"):
                    os.startfile(out_path)
            except Exception as e:
                messagebox.showerror("خطا", str(e))

        btn_row = ttk.Frame(parent)
        btn_row.pack(pady=10)
        ttk.Button(btn_row, text="بروزرسانی گزارش‌ها", command=refresh).pack(side="right", padx=5)
        ttk.Button(btn_row, text="خروجی اکسل بدهکاران/بستانکاران", command=export_all_excel).pack(side="right", padx=5)
        refresh()
        self.tab_refresh_callbacks[parent] = refresh

    # ---------------- مدیریت کاربران (فقط مدیر) ----------------
    def build_users_tab(self, parent):
        top = ttk.Frame(parent, padding=10)
        top.pack(fill="x")

        ttk.Label(top, text="نام کاربری:").grid(row=0, column=3, padx=5)
        user_e = ttk.Entry(top, width=15)
        user_e.grid(row=0, column=2, padx=5)

        ttk.Label(top, text="رمز عبور:").grid(row=0, column=1, padx=5)
        pass_e = ttk.Entry(top, width=15, show="*")
        pass_e.grid(row=0, column=0, padx=5)

        ttk.Label(top, text="سطح دسترسی:").grid(row=1, column=1, padx=5)
        role_combo = ttk.Combobox(top, width=12, state="readonly", values=["admin", "employee"])
        role_combo.set("employee")
        role_combo.grid(row=1, column=0, padx=5)

        tree = ttk.Treeview(parent, columns=("id", "username", "role"), show="headings")
        for c, t in zip(("id", "username", "role"), ("شماره", "نام کاربری", "سطح دسترسی")):
            tree.heading(c, text=t)
            tree.column(c, width=150, anchor="center")
        tree.pack(fill="both", expand=True, padx=10, pady=10)

        def refresh():
            for r in tree.get_children():
                tree.delete(r)
            users = self.api("GET", "/users") or []
            for u in users:
                tree.insert("", "end", values=(u["id"], u["username"],
                                                "مدیر" if u["role"] == "admin" else "کارمند"))

        def add_user():
            if not user_e.get() or not pass_e.get():
                messagebox.showerror("خطا", "نام کاربری و رمز عبور الزامی است")
                return
            res = self.api("POST", "/users", json={
                "username": user_e.get(), "password": pass_e.get(), "role": role_combo.get()
            })
            if res:
                if res.get("ok"):
                    messagebox.showinfo("موفق", res.get("message", "کاربر اضافه شد"))
                    refresh()
                else:
                    messagebox.showerror("خطا", res.get("message", "خطا در افزودن کاربر"))

        def delete_user():
            sel = tree.selection()
            if not sel:
                messagebox.showerror("خطا", "یک کاربر را انتخاب کنید")
                return
            uid = tree.item(sel[0])["values"][0]
            if uid == self.user["id"]:
                messagebox.showerror("خطا", "نمی‌توانید کاربر جاری خودتان را حذف کنید")
                return
            if messagebox.askyesno("تأیید", "آیا از حذف این کاربر مطمئن هستید؟"):
                res = self.api("DELETE", f"/users/{uid}")
                if res and res.get("ok"):
                    refresh()

        ttk.Button(top, text="افزودن کاربر", command=add_user).grid(row=0, column=4, padx=10)
        ttk.Button(parent, text="حذف کاربر انتخاب‌شده", command=delete_user).pack(pady=5)
        ttk.Label(parent, text="توجه: فقط کاربران با سطح دسترسی «مدیر» این تب‌ها (پشتیبان‌گیری، مدیریت کاربران) را می‌بینند.",
                  foreground="gray").pack(pady=5)
        refresh()
        self.tab_refresh_callbacks[parent] = refresh

    # ---------------- لاگ فعالیت کاربران (فقط مدیر) ----------------
    def build_activity_log_tab(self, parent):
        top = ttk.Frame(parent, padding=10)
        top.pack(fill="x")

        tree = ttk.Treeview(parent, columns=("time", "user", "action", "details"), show="headings")
        for c, t in zip(("time", "user", "action", "details"), ("زمان", "کاربر", "فعالیت", "جزئیات")):
            tree.heading(c, text=t)
            tree.column(c, width=150 if c != "details" else 220, anchor="center")
        tree.pack(fill="both", expand=True, padx=10, pady=10)

        def refresh():
            for r in tree.get_children():
                tree.delete(r)
            logs = self.api("GET", "/activity-log") or []
            for lg in logs:
                tree.insert("", "end", values=(lg["timestamp"], lg["username"], lg["action"], lg["details"]))

        ttk.Button(top, text="بروزرسانی لاگ", command=refresh).pack(pady=5)
        refresh()
        self.tab_refresh_callbacks[parent] = refresh


if __name__ == "__main__":
    app = App()
    app.mainloop()
