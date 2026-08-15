# -*- coding: utf-8 -*-
"""
سرور حسابداری مغازه الکتریکی
این فایل روی کامپیوتر اصلی (سرور) اجرا میشه و روی شبکه داخلی (LAN) گوش میده.
کاملاً آفلاین کار می‌کنه - هیچ نیازی به اینترنت نداره.
"""
from flask import Flask, request, jsonify, send_file, send_from_directory, g
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from database import init_db, get_connection, DB_PATH
from datetime import datetime, timedelta
import os
import random
import secrets
import sqlite3
import shutil
import threading
import time
import tempfile
import json
import notifier
import pdf_generator
import invoice_html
import invoice_ai
import invoice_ocr_free
import assistant_ai
from paths import get_base_dir, get_bundle_dir

SETTINGS_PATH = os.path.join(get_base_dir(), "shop_settings.json")
ASSETS_DIR = os.path.join(get_base_dir(), "assets")

# اطمینان از وجود پوشه assets و آیکون پیش‌فرض کنار فایل exe (چون پوشه‌ی داخل exe موقتی است)
os.makedirs(ASSETS_DIR, exist_ok=True)
_default_icon_src = os.path.join(get_bundle_dir(), "assets", "icon.png")
_default_icon_dst = os.path.join(ASSETS_DIR, "icon.png")
if os.path.exists(_default_icon_src) and not os.path.exists(_default_icon_dst):
    shutil.copy2(_default_icon_src, _default_icon_dst)


def load_shop_settings():
    if os.path.exists(SETTINGS_PATH):
        try:
            with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"name": "حسابداری", "phones": "", "address": "", "logo_filename": None, "invoice_number_offset": 0,
            "default_margin_percent": 0, "invoice_footer_message": "",
            "national_id": "", "economic_code": "", "postal_code": "",
            "ai_enabled": False, "ai_api_key": ""}


def get_next_invoice_id(conn):
    """
    شماره‌ی بعدی که SQLite برای ستون AUTOINCREMENT جدول invoices اختصاص می‌دهد.
    از sqlite_sequence استفاده می‌شود (نه MAX(id)) چون AUTOINCREMENT حتی بعد از حذف
    فاکتورها هم دوباره از یک شماره‌ی قبلی استفاده نمی‌کند.
    """
    row = conn.execute("SELECT seq FROM sqlite_sequence WHERE name='invoices'").fetchone()
    last_seq = row["seq"] if row else 0
    return last_seq + 1


def save_shop_settings(data):
    with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def notify_telegram_async(text):
    """ارسال پیام هشدار به تلگرام صاحب مغازه در یک ترد جدا، تا کند بودن اینترنت
    باعث معطل ماندن درخواست اصلی (مثلاً ثبت فاکتور) نشود."""
    cfg = load_shop_settings()
    token = cfg.get("telegram_bot_token")
    chat_id = cfg.get("telegram_chat_id")
    if not token or not chat_id:
        return
    threading.Thread(target=notifier.send_telegram_message, args=(token, chat_id, text), daemon=True).start()


WEB_DIR = os.path.join(get_bundle_dir(), "web")
app = Flask(__name__, static_folder=WEB_DIR, static_url_path="")


@app.after_request
def add_no_cache_headers(response):
    """جلوگیری از کش کردن صفحه توسط مرورگر، تا هر تغییری فوراً دیده بشه"""
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return response


# ---------- احراز هویت (Session Token) ----------
# چون سرور روی شبکه (LAN) گوش می‌ده، هر درخواستی (به‌جز لاگین و فایل‌های عمومی) باید
# یک توکن معتبر همراه داشته باشه، وگرنه هر دستگاهی تو شبکه بدون لاگین به همه‌ی اطلاعات
# مالی و مدیریتی دسترسی پیدا می‌کرد.
SESSIONS = {}
SESSIONS_LOCK = threading.Lock()

# مسیرهایی که قبل از لاگین هم باید در دسترس باشند (صفحه ورود، فایل‌های استاتیک آن، پینگ)
PUBLIC_ENDPOINTS = {"static", "index", "login", "logout", "ping", "assets", "get_shop_settings", "get_captcha"}

# ---------- کد امنیتی (کپچا) و قفل موقت بعد از تلاش‌های ناموفق ورود ----------
CAPTCHAS = {}
CAPTCHA_LOCK = threading.Lock()
CAPTCHA_TTL_SECONDS = 300  # هر کد امنیتی حداکثر ۵ دقیقه معتبر است

FAILED_LOGINS = {}  # username (lowercase) -> {"count": int, "locked_until": timestamp}
FAILED_LOGINS_LOCK = threading.Lock()
MAX_LOGIN_ATTEMPTS = 5
LOCKOUT_SECONDS = 300  # قفل ۵ دقیقه‌ای بعد از ۵ تلاش ناموفق پشت سر هم


def generate_captcha():
    a, b = random.randint(1, 9), random.randint(1, 9)
    op = random.choice(["+", "-"])
    answer = a + b if op == "+" else a - b
    captcha_id = secrets.token_hex(16)
    now_ts = time.time()
    with CAPTCHA_LOCK:
        expired = [cid for cid, v in CAPTCHAS.items() if v["expires_at"] < now_ts]
        for cid in expired:
            CAPTCHAS.pop(cid, None)
        CAPTCHAS[captcha_id] = {"answer": answer, "expires_at": now_ts + CAPTCHA_TTL_SECONDS}
    return captcha_id, f"{a} {op} {b} = ?"


def verify_and_consume_captcha(captcha_id, answer):
    """هر کد امنیتی فقط یک‌بار قابل استفاده است (بلافاصله بعد از بررسی حذف می‌شود)"""
    if not captcha_id:
        return False
    with CAPTCHA_LOCK:
        data = CAPTCHAS.pop(captcha_id, None)
    if not data or data["expires_at"] < time.time():
        return False
    try:
        return int(answer) == data["answer"]
    except (TypeError, ValueError):
        return False


# ---------- ورود دو مرحله‌ای (TOTP سازگار با Google Authenticator/Authy و مشابه) ----------
import base64
import hmac
import hashlib
import struct


def generate_totp_secret():
    return base64.b32encode(secrets.token_bytes(10)).decode("utf-8")


def totp_code_at(secret, for_time, interval=30, digits=6):
    padded = secret.upper() + "=" * ((8 - len(secret) % 8) % 8)
    key = base64.b32decode(padded)
    counter = int(for_time // interval)
    msg = struct.pack(">Q", counter)
    h = hmac.new(key, msg, hashlib.sha1).digest()
    offset = h[-1] & 0x0F
    code_int = (struct.unpack(">I", h[offset:offset + 4])[0] & 0x7FFFFFFF) % (10 ** digits)
    return str(code_int).zfill(digits)


def verify_totp(secret, code, window=1):
    """کد را با یک بازه‌ی ±۳۰ ثانیه بررسی می‌کند تا اختلاف جزئی ساعت گوشی مشکلی ایجاد نکند"""
    if not secret or not code:
        return False
    code = str(code).strip()
    now_ts = time.time()
    for i in range(-window, window + 1):
        if hmac.compare_digest(totp_code_at(secret, now_ts + i * 30), code):
            return True
    return False


# ---------- سطح دسترسی سفارشی کارمندها (فراتر از مدیر/کارمند ساده) ----------
PERMISSION_KEYS = ("can_sell", "can_purchase", "can_manage_items", "can_manage_parties", "can_manage_cash")


def get_user_permissions(username):
    conn = get_connection()
    row = conn.execute("SELECT permissions FROM users WHERE username=?", (username,)).fetchone()
    conn.close()
    if not row or not row["permissions"]:
        return {}
    try:
        return json.loads(row["permissions"])
    except (TypeError, ValueError):
        return {}


def require_permission(key):
    """اگر کاربر ادمین نباشد و صراحتاً این دسترسی از او گرفته شده باشد (پیش‌فرض همه چیز مجاز است
    مگر مدیر آن را خاموش کرده باشد)، خطای ۴۰۳ برمی‌گرداند."""
    if g.current_user.get("role") == "admin":
        return None
    perms = get_user_permissions(g.current_user["username"])
    if perms.get(key, True) is False:
        return jsonify({"ok": False, "message": "شما اجازه دسترسی به این بخش را ندارید — از مدیر بخواهید دسترسی لازم را فعال کند"}), 403
    return None


def create_session(username, role):
    token = secrets.token_hex(32)
    with SESSIONS_LOCK:
        SESSIONS[token] = {"username": username, "role": role}
    return token


def get_current_session():
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        return None
    token = auth_header[len("Bearer "):].strip()
    with SESSIONS_LOCK:
        return SESSIONS.get(token)


@app.before_request
def require_auth():
    if request.method == "OPTIONS" or request.endpoint in PUBLIC_ENDPOINTS:
        return None
    session = get_current_session()
    if not session:
        return jsonify({"ok": False, "message": "نیاز به ورود مجدد به سیستم است"}), 401
    g.current_user = session


def require_admin():
    """برای مسیرهای مخصوص مدیر؛ اگر کاربر ادمین نباشد پاسخ خطا برمی‌گرداند"""
    if g.current_user.get("role") != "admin":
        return jsonify({"ok": False, "message": "این عملیات فقط برای مدیر سیستم مجاز است"}), 403
    return None


@app.route("/settings/shop", methods=["GET"])
def get_shop_settings():
    s = load_shop_settings()
    s["logo_url"] = f"/assets/{s['logo_filename']}" if s.get("logo_filename") else None
    conn = get_connection()
    next_id = get_next_invoice_id(conn)
    conn.close()
    s["next_invoice_number"] = next_id + (s.get("invoice_number_offset", 0) or 0)
    # این مسیر حتی قبل از ورود هم صدا زده می‌شود (برای نمایش نام/لوگوی مغازه در صفحه ورود)،
    # پس کلیدهای حساس هرگز نباید در پاسخش برگردند — فقط این‌که «تنظیم شده یا نه»
    s["telegram_bot_token_set"] = bool(s.pop("telegram_bot_token", None))
    s["ai_api_key_set"] = bool(s.pop("ai_api_key", None))
    return jsonify(s)


@app.route("/settings/shop", methods=["POST"])
def update_shop_settings():
    err = require_admin()
    if err:
        return err
    d = request.json
    s = load_shop_settings()
    s["name"] = d.get("name", s.get("name"))
    s["phones"] = d.get("phones", s.get("phones"))
    s["address"] = d.get("address", s.get("address"))
    if d.get("telegram_bot_token"):
        s["telegram_bot_token"] = d.get("telegram_bot_token").strip()
    if "telegram_chat_id" in d:
        s["telegram_chat_id"] = d.get("telegram_chat_id") or ""
    if "ai_enabled" in d:
        s["ai_enabled"] = bool(d.get("ai_enabled"))
    if d.get("ai_api_key"):
        s["ai_api_key"] = d.get("ai_api_key").strip()
    if "invoice_footer_message" in d:
        s["invoice_footer_message"] = d.get("invoice_footer_message") or ""
    if "national_id" in d:
        s["national_id"] = d.get("national_id") or ""
    if "economic_code" in d:
        s["economic_code"] = d.get("economic_code") or ""
    if "postal_code" in d:
        s["postal_code"] = d.get("postal_code") or ""
    if "default_margin_percent" in d:
        try:
            s["default_margin_percent"] = float(d.get("default_margin_percent") or 0)
        except (TypeError, ValueError):
            return jsonify({"ok": False, "message": "درصد سود باید عدد باشد"}), 400
    if d.get("next_invoice_number") not in (None, ""):
        try:
            next_number = int(d["next_invoice_number"])
        except (TypeError, ValueError):
            return jsonify({"ok": False, "message": "شماره فاکتور بعدی باید عدد باشد"}), 400
        if next_number < 1:
            return jsonify({"ok": False, "message": "شماره فاکتور بعدی باید حداقل ۱ باشد"}), 400
        conn = get_connection()
        next_id = get_next_invoice_id(conn)
        conn.close()
        s["invoice_number_offset"] = next_number - next_id
    save_shop_settings(s)
    return jsonify({"ok": True})


@app.route("/settings/telegram/test", methods=["POST"])
def test_telegram():
    err = require_admin()
    if err:
        return err
    s = load_shop_settings()
    ok, message = notifier.send_telegram_message(
        s.get("telegram_bot_token"), s.get("telegram_chat_id"),
        "✅ این یک پیام تستی از برنامه حسابداری مغازه شماست."
    )
    return jsonify({"ok": ok, "message": message})


@app.route("/settings/shop/logo", methods=["POST"])
def upload_shop_logo():
    err = require_admin()
    if err:
        return err
    if "logo" not in request.files:
        return jsonify({"ok": False, "message": "فایلی ارسال نشده"}), 400
    file = request.files["logo"]
    if not file or file.filename == "":
        return jsonify({"ok": False, "message": "فایلی انتخاب نشده"}), 400
    ext = os.path.splitext(file.filename)[1].lower()
    # svg عمداً مجاز نیست: چون این فایل مستقیماً به‌عنوان یک فایل استاتیک سرو می‌شود،
    # svg با اسکریپت داخلش می‌تواند حمله XSS ذخیره‌شده ایجاد کند
    if ext not in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
        return jsonify({"ok": False, "message": "فرمت فایل باید png، jpg، gif یا webp باشد"}), 400
    os.makedirs(ASSETS_DIR, exist_ok=True)
    filename = "shop_logo" + ext
    file.save(os.path.join(ASSETS_DIR, filename))
    s = load_shop_settings()
    s["logo_filename"] = filename
    save_shop_settings(s)
    return jsonify({"ok": True, "logo_url": f"/assets/{filename}"})


@app.route("/settings/shop/logo", methods=["DELETE"])
def delete_shop_logo():
    err = require_admin()
    if err:
        return err
    s = load_shop_settings()
    old = s.get("logo_filename")
    if old:
        old_path = os.path.join(ASSETS_DIR, old)
        if os.path.exists(old_path):
            os.remove(old_path)
    s["logo_filename"] = None
    save_shop_settings(s)
    return jsonify({"ok": True})


BACKUP_DIR = os.path.join(get_base_dir(), "backups")
os.makedirs(BACKUP_DIR, exist_ok=True)


# ---------- سرو کردن رابط کاربری وب ----------
@app.route("/")
def index():
    return send_from_directory(WEB_DIR, "index.html")


@app.route("/assets/<path:filename>")
def assets(filename):
    return send_from_directory(os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets"), filename)


def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def log_action(username, action, details=""):
    try:
        conn = get_connection()
        conn.execute("INSERT INTO activity_log (timestamp, username, action, details) VALUES (?,?,?,?)",
                     (now(), username or "نامشخص", action, details))
        conn.commit()
        conn.close()
    except Exception:
        pass


def log_security_event(username, event, details=""):
    try:
        conn = get_connection()
        conn.execute("INSERT INTO security_log (timestamp, username, event, details) VALUES (?,?,?,?)",
                     (now(), username or "نامشخص", event, details))
        conn.commit()
        conn.close()
    except Exception:
        pass


# ---------- ورود ----------
@app.route("/captcha", methods=["GET"])
def get_captcha():
    captcha_id, question = generate_captcha()
    return jsonify({"captcha_id": captcha_id, "question": question})


@app.route("/login", methods=["POST"])
def login():
    data = request.json or {}
    username = (data.get("username") or "").strip()
    username_key = username.lower()

    # ورود اضطراری: اگر کنار accounting.db فایل خالی RESET_ADMIN_PASSWORD.txt وجود داشته باشد،
    # صرفاً با زدن نام کاربری admin (بدون نیاز به رمز یا کد امنیتی درست) وارد می‌شود — چون تایپ
    # دقیق رمز/کد امنیتی برای بعضی کاربران مشکل‌ساز شده بود. بلافاصله بعد از ورود مجبور به تعیین
    # رمز جدید می‌شود. این راه فقط با دسترسی فیزیکی به کامپیوتر (ساختن آن فایل) ممکن است.
    emergency_marker_path = os.path.join(get_base_dir(), "RESET_ADMIN_PASSWORD.txt")
    if username_key == "admin" and os.path.exists(emergency_marker_path):
        conn = get_connection()
        admin_user = conn.execute("SELECT * FROM users WHERE username='admin'").fetchone()
        if admin_user:
            conn.execute("UPDATE users SET must_change_password=1 WHERE username='admin'")
            conn.commit()
            user_dict = dict(admin_user)
            user_dict.pop("password", None)
            user_dict.pop("totp_secret", None)
            user_dict["must_change_password"] = 1
            token = create_session("admin", admin_user["role"])
            conn.close()
            try:
                os.remove(emergency_marker_path)
            except OSError:
                pass
            with FAILED_LOGINS_LOCK:
                FAILED_LOGINS.pop(username_key, None)
            log_action("admin", "ورود اضطراری با فایل بازیابی")
            return jsonify({"ok": True, "user": user_dict, "token": token})
        conn.close()

    with FAILED_LOGINS_LOCK:
        rec = FAILED_LOGINS.get(username_key)
        if rec and rec.get("locked_until", 0) > time.time():
            remaining = int(rec["locked_until"] - time.time()) + 1
            return jsonify({
                "ok": False,
                "message": f"به‌خاطر تلاش‌های ناموفق زیاد، این حساب موقتاً قفل شده. {remaining} ثانیه دیگر دوباره امتحان کنید.",
                "retry_after_seconds": remaining,
            }), 429

    if not verify_and_consume_captcha(data.get("captcha_id"), data.get("captcha_answer")):
        return jsonify({"ok": False, "message": "کد امنیتی اشتباه است یا منقضی شده", "captcha_error": True}), 400

    conn = get_connection()
    user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    conn.close()
    if user and check_password_hash(user["password"], data.get("password", "")):
        if user["totp_enabled"]:
            totp_code = (data.get("totp_code") or "").strip()
            if not totp_code:
                return jsonify({"ok": False, "need_totp": True, "message": "کد ورود دو مرحله‌ای را وارد کنید"}), 200
            if not verify_totp(user["totp_secret"], totp_code):
                log_security_event(user["username"], "ورود ناموفق", "کد دو مرحله‌ای اشتباه")
                return jsonify({"ok": False, "need_totp": True, "message": "کد دو مرحله‌ای اشتباه است"}), 401
        with FAILED_LOGINS_LOCK:
            FAILED_LOGINS.pop(username_key, None)
        log_action(user["username"], "ورود به سیستم")
        user_dict = dict(user)
        user_dict.pop("password", None)  # رمز هش‌شده هم نباید به کلاینت فرستاده بشه
        user_dict.pop("totp_secret", None)
        token = create_session(user["username"], user["role"])
        return jsonify({"ok": True, "user": user_dict, "token": token})

    if username_key:
        with FAILED_LOGINS_LOCK:
            rec = FAILED_LOGINS.setdefault(username_key, {"count": 0, "locked_until": 0})
            rec["count"] += 1
            if rec["count"] >= MAX_LOGIN_ATTEMPTS:
                rec["locked_until"] = time.time() + LOCKOUT_SECONDS
                rec["count"] = 0
    log_security_event(username, "ورود ناموفق", "نام کاربری یا رمز اشتباه")
    return jsonify({"ok": False, "message": "نام کاربری یا رمز عبور اشتباه است"}), 401


@app.route("/logout", methods=["POST"])
def logout():
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        token = auth_header[len("Bearer "):].strip()
        with SESSIONS_LOCK:
            SESSIONS.pop(token, None)
    return jsonify({"ok": True})


# ---------- کالاها ----------
@app.route("/items", methods=["GET"])
def get_items():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM items WHERE deleted_at IS NULL ORDER BY name").fetchall()
    conn.close()
    result = [dict(r) for r in rows]
    if g.current_user.get("role") != "admin":
        for r in result:
            r["purchase_price"] = None
    return jsonify(result)


@app.route("/items", methods=["POST"])
def add_item():
    err = require_permission("can_manage_items")
    if err:
        return err
    d = request.json
    conn = get_connection()
    try:
        conn.execute("""INSERT INTO items (code, name, category_id, unit, purchase_price, sale_price, stock_qty, min_stock, brand)
                         VALUES (?,?,?,?,?,?,?,?,?)""",
                     (d.get("code") or None, d["name"], d.get("category_id"), d.get("unit", "عدد"),
                      d.get("purchase_price", 0), d.get("sale_price", 0),
                      d.get("stock_qty", 0), d.get("min_stock", 0), d.get("brand")))
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"ok": False, "message": "کالای دیگری با همین کد/بارکد قبلاً ثبت شده"}), 400
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/items/<int:item_id>", methods=["PUT"])
def update_item(item_id):
    err = require_permission("can_manage_items")
    if err:
        return err
    d = request.json
    conn = get_connection()
    try:
        conn.execute("""UPDATE items SET code=?, name=?, category_id=?, unit=?, purchase_price=?,
                         sale_price=?, stock_qty=?, min_stock=?, brand=? WHERE id=?""",
                     (d.get("code") or None, d["name"], d.get("category_id"), d.get("unit", "عدد"),
                      d.get("purchase_price", 0), d.get("sale_price", 0),
                      d.get("stock_qty", 0), d.get("min_stock", 0), d.get("brand"), item_id))
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"ok": False, "message": "کالای دیگری با همین کد/بارکد قبلاً ثبت شده"}), 400
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/items/<int:item_id>", methods=["DELETE"])
def delete_item(item_id):
    """حذف نرم: کالا از لیست‌های فعال کنار می‌رود ولی تا وقتی از سطل زباله برای همیشه
    حذف نشود، قابل بازگردانی است — تا اشتباهی برای همیشه از دست نرود"""
    err = require_permission("can_manage_items")
    if err:
        return err
    conn = get_connection()
    conn.execute("UPDATE items SET deleted_at=? WHERE id=?", (now(), item_id))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/items/trash", methods=["GET"])
def get_items_trash():
    err = require_permission("can_manage_items")
    if err:
        return err
    conn = get_connection()
    rows = conn.execute("SELECT * FROM items WHERE deleted_at IS NOT NULL ORDER BY deleted_at DESC").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/items/<int:item_id>/restore", methods=["POST"])
def restore_item(item_id):
    err = require_permission("can_manage_items")
    if err:
        return err
    conn = get_connection()
    conn.execute("UPDATE items SET deleted_at=NULL WHERE id=?", (item_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/items/<int:item_id>/permanent", methods=["DELETE"])
def permanently_delete_item(item_id):
    """حذف همیشگی از سطل زباله — این یکی واقعاً قابل بازگشت نیست"""
    err = require_admin()
    if err:
        return err
    conn = get_connection()
    conn.execute("DELETE FROM items WHERE id=? AND deleted_at IS NOT NULL", (item_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/items/<int:item_id>/photo", methods=["POST"])
def upload_item_photo(item_id):
    if "photo" not in request.files:
        return jsonify({"ok": False, "message": "فایلی ارسال نشده"}), 400
    file = request.files["photo"]
    if not file or file.filename == "":
        return jsonify({"ok": False, "message": "فایلی انتخاب نشده"}), 400
    ext = os.path.splitext(file.filename)[1].lower()
    # svg عمداً مجاز نیست: چون این فایل مستقیماً به‌عنوان یک فایل استاتیک سرو می‌شود،
    # svg با اسکریپت داخلش می‌تواند حمله XSS ذخیره‌شده ایجاد کند
    if ext not in (".png", ".jpg", ".jpeg", ".gif", ".webp"):
        return jsonify({"ok": False, "message": "فرمت فایل باید png، jpg، gif یا webp باشد"}), 400
    conn = get_connection()
    item = conn.execute("SELECT * FROM items WHERE id=?", (item_id,)).fetchone()
    if not item:
        conn.close()
        return jsonify({"ok": False, "message": "کالا پیدا نشد"}), 404
    os.makedirs(ASSETS_DIR, exist_ok=True)
    filename = f"item_{item_id}_{secrets.token_hex(4)}{ext}"
    old_photo = item["photo_filename"]
    file.save(os.path.join(ASSETS_DIR, filename))
    conn.execute("UPDATE items SET photo_filename=? WHERE id=?", (filename, item_id))
    conn.commit()
    conn.close()
    if old_photo:
        old_path = os.path.join(ASSETS_DIR, old_photo)
        if os.path.exists(old_path):
            os.remove(old_path)
    return jsonify({"ok": True, "photo_url": f"/assets/{filename}"})


# ---------- طرف‌حساب‌ها (مشتری/تامین‌کننده) ----------
@app.route("/parties", methods=["GET"])
def get_parties():
    ptype = request.args.get("type")
    conn = get_connection()
    if ptype:
        rows = conn.execute("SELECT * FROM parties WHERE type=? ORDER BY name", (ptype,)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM parties ORDER BY name").fetchall()
    result = [dict(r) for r in rows]
    for p in result:
        last = conn.execute(
            "SELECT MAX(date) as d FROM invoices WHERE party_id=? AND invoice_type IN ('sale','purchase')",
            (p["id"],)
        ).fetchone()
        p["last_purchase"] = last["d"] if last else None
    conn.close()
    return jsonify(result)


import re

PHONE_PATTERN = re.compile(r"^(0?9\d{9}|989\d{9})$")


def validate_party_input(d):
    """اعتبارسنجی شماره تماس (۱۱ یا ۱۲ رقم) و اجباری بودن آدرس"""
    phone = (d.get("phone") or "").strip().replace(" ", "").replace("-", "")
    if not phone or not PHONE_PATTERN.match(phone):
        return "شماره تماس نامعتبر است — باید ۱۱ رقم (مثلاً 09123456789) یا ۱۲ رقم با پیش‌شماره کشور باشد"
    if not (d.get("address") or "").strip():
        return "آدرس محل کار یا سکونت الزامی است"
    return None


@app.route("/parties", methods=["POST"])
def add_party():
    d = request.json
    if not d.get("_auto_created"):
        err = require_permission("can_manage_parties")
        if err:
            return err
        error = validate_party_input(d)
        if error:
            return jsonify({"ok": False, "message": error}), 400
    conn = get_connection()
    cur = conn.execute("""INSERT INTO parties (name, phone, address, type, balance, note, credit_limit, is_vip, special_discount_percent)
                     VALUES (?,?,?,?,?,?,?,?,?)""",
                 (d["name"], d.get("phone"), d.get("address"), d.get("type", "customer"), d.get("balance", 0),
                  d.get("note", ""), d.get("credit_limit", 0) or 0, 1 if d.get("is_vip") else 0,
                  d.get("special_discount_percent", 0) or 0))
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return jsonify({"ok": True, "id": new_id})


@app.route("/parties/<int:party_id>", methods=["PUT"])
def update_party(party_id):
    err = require_permission("can_manage_parties")
    if err:
        return err
    d = request.json
    conn = get_connection()
    conn.execute(
        "UPDATE parties SET name=?, phone=?, address=?, note=?, credit_limit=?, is_vip=?, special_discount_percent=? WHERE id=?",
        (d["name"], d.get("phone"), d.get("address"), d.get("note", ""), d.get("credit_limit", 0) or 0,
         1 if d.get("is_vip") else 0, d.get("special_discount_percent", 0) or 0, party_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/parties/<int:party_id>/payment", methods=["POST"])
def party_payment(party_id):
    """
    ثبت تسویه حساب: دریافت وجه از مشتری بدهکار، یا پرداخت به تامین‌کننده.
    ورودی: {"amount": مبلغ, "description": "..."}
    """
    d = request.json
    amount = d["amount"]
    conn = get_connection()
    party = conn.execute("SELECT * FROM parties WHERE id=?", (party_id,)).fetchone()
    if not party:
        conn.close()
        return jsonify({"ok": False, "message": "طرف حساب پیدا نشد"}), 404

    if party["type"] == "customer":
        new_balance = party["balance"] - amount   # دریافت وجه، بدهی مشتری کم می‌شود
        tx_type = "in"
    else:
        new_balance = party["balance"] + amount   # پرداخت به تامین‌کننده
        tx_type = "out"

    conn.execute("UPDATE parties SET balance=? WHERE id=?", (new_balance, party_id))
    conn.execute("INSERT INTO cash_transactions (date, tx_type, amount, description) VALUES (?,?,?,?)",
                 (now(), tx_type, amount, d.get("description") or f"تسویه حساب با {party['name']}"))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "new_balance": new_balance})


@app.route("/parties/<int:party_id>/ledger", methods=["GET"])
def party_ledger(party_id):
    """ریز فاکتورهای یک طرف‌حساب به‌همراه مانده فعلی"""
    conn = get_connection()
    party = conn.execute("SELECT * FROM parties WHERE id=?", (party_id,)).fetchone()
    if not party:
        conn.close()
        return jsonify({"ok": False, "message": "طرف حساب پیدا نشد"}), 404
    invoices = conn.execute(
        "SELECT * FROM invoices WHERE party_id=? AND voided=0 ORDER BY date", (party_id,)
    ).fetchall()
    conn.close()
    return jsonify({"party": dict(party), "invoices": [dict(r) for r in invoices]})


# ---------- دسته‌بندی کالا ----------
@app.route("/categories", methods=["GET"])
def get_categories():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM categories ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/categories", methods=["POST"])
def add_category():
    d = request.json
    conn = get_connection()
    try:
        conn.execute("INSERT INTO categories (name) VALUES (?)", (d["name"],))
        conn.commit()
        ok, msg = True, "دسته اضافه شد"
    except Exception as e:
        ok = False
        msg = "این دسته قبلاً وجود دارد" if "UNIQUE" in str(e) else str(e)
    conn.close()
    return jsonify({"ok": ok, "message": msg})


# ---------- فاکتورها (فروش / خرید) ----------
@app.route("/invoices", methods=["GET"])
def get_invoices():
    itype = request.args.get("type")
    conn = get_connection()
    q = "SELECT invoices.*, parties.name as party_name FROM invoices LEFT JOIN parties ON invoices.party_id = parties.id"
    params = ()
    if itype:
        q += " WHERE invoice_type=?"
        params = (itype,)
    q += " ORDER BY date DESC, invoices.id DESC"
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/invoices/last-purchase/<int:party_id>", methods=["GET"])
def get_last_purchase_invoice(party_id):
    """آخرین فاکتور خرید ثبت‌شده از یک تامین‌کننده، برای دکمه‌ی «تکرار آخرین فاکتور»"""
    conn = get_connection()
    inv = conn.execute(
        "SELECT * FROM invoices WHERE party_id=? AND invoice_type='purchase' AND voided=0 ORDER BY date DESC, id DESC LIMIT 1",
        (party_id,)
    ).fetchone()
    if not inv:
        conn.close()
        return jsonify({"ok": False, "message": "فاکتور خریدی از این تامین‌کننده ثبت نشده"}), 404
    items = conn.execute(
        "SELECT item_id, qty, unit_price FROM invoice_items WHERE invoice_id=?", (inv["id"],)
    ).fetchall()
    conn.close()
    return jsonify({"ok": True, "invoice_number": inv["number"], "items": [dict(r) for r in items]})


@app.route("/invoices/<int:inv_id>/items", methods=["GET"])
def get_invoice_items(inv_id):
    conn = get_connection()
    rows = conn.execute("""SELECT invoice_items.*, items.name as item_name FROM invoice_items
                            JOIN items ON invoice_items.item_id = items.id
                            WHERE invoice_id=?""", (inv_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/invoices", methods=["POST"])
def add_invoice():
    """
    ساختار ورودی:
    {
      "invoice_type": "sale" | "purchase" | "sale_return" | "purchase_return",
      "party_id": int یا null,
      "payment_type": "cash" | "credit" | "check",
      "discount": مبلغ تخفیف (اختیاری، پیش‌فرض ۰),
      "description": "...",
      "username": "برای لاگ فعالیت (اختیاری)",
      "items": [{"item_id": 1, "qty": 2, "unit_price": 15000}, ...]
    }
    مرجوعی خرید/فروش موجودی انبار را برعکسِ خرید/فروش عادی تغییر می‌دهد.
    """
    d = request.json
    invoice_type = d["invoice_type"]
    perm_key = "can_sell" if invoice_type in ("sale", "purchase_return") else "can_purchase"
    err = require_permission(perm_key)
    if err:
        return err

    conn = get_connection()
    c = conn.cursor()

    is_return = invoice_type in ("sale_return", "purchase_return")
    discount = d.get("discount", 0) or 0
    subtotal = sum(it["qty"] * it["unit_price"] for it in d["items"])
    total = max(subtotal - discount, 0)
    paid = total if d.get("payment_type") == "cash" else d.get("paid", 0)

    # اعتبارسنجی سقف اعتبار نسیه (اگر برای مشتری تعیین شده باشد)
    if invoice_type == "sale" and d.get("payment_type") == "credit" and d.get("party_id"):
        party = conn.execute("SELECT * FROM parties WHERE id=?", (d["party_id"],)).fetchone()
        if party and party["credit_limit"] and party["credit_limit"] > 0:
            remaining = total - paid
            projected_balance = party["balance"] + remaining
            if projected_balance > party["credit_limit"]:
                conn.close()
                return jsonify({
                    "ok": False,
                    "message": f'این فاکتور از سقف اعتبار «{party["name"]}» عبور می‌کند '
                               f'(سقف: {party["credit_limit"]:,.0f}، بدهی فعلی: {party["balance"]:,.0f}، بعد از این فاکتور: {projected_balance:,.0f})'
                }), 400

    # اعتبارسنجی موجودی کافی برای فاکتور فروش (نباید موجودی منفی بشه)
    if invoice_type == "sale":
        needed = {}
        for it in d["items"]:
            needed[it["item_id"]] = needed.get(it["item_id"], 0) + it["qty"]
        for item_id, qty_needed in needed.items():
            row = conn.execute("SELECT name, stock_qty FROM items WHERE id=?", (item_id,)).fetchone()
            if row and row["stock_qty"] < qty_needed:
                conn.close()
                return jsonify({
                    "ok": False,
                    "message": f'موجودی «{row["name"]}» فقط {row["stock_qty"]:g} عدد است — نمی‌توان {qty_needed:g} تا فروخت'
                }), 400

    # شماره فاکتور خودکار: بعد از درج، بر اساس id ساخته می‌شود.
    # فاکتور فروش عمداً بدون پیشوند حرفی است (فقط عدد ساده، بدون صفرهای اضافه جلوش) —
    # چون این همون شماره‌ایه که مشتری روی فاکتورش می‌بینه. بقیه‌ی انواع فاکتور (خرید/مرجوعی)
    # برای تشخیص از هم، پیشوند حرفی و padding پنج‌رقمی دارند.
    prefix_map = {"sale": "", "purchase": "PU", "sale_return": "SR", "purchase_return": "PR"}
    prefix = prefix_map.get(invoice_type, "IN")

    c.execute("""INSERT INTO invoices (invoice_type, number, party_id, date, total, paid, payment_type, description, discount, created_by)
                 VALUES (?,?,?,?,?,?,?,?,?,?)""",
              (invoice_type, None, d.get("party_id"), now(),
               total, paid, d.get("payment_type", "cash"), d.get("description", ""), discount,
               g.current_user["username"]))
    invoice_id = c.lastrowid
    # اگر از تنظیمات یه «شماره فاکتور بعدی» دلخواه تنظیم شده باشه (مثلاً برای ادامه‌ی شماره‌گذاری
    # فاکتورهای کاغذی قبلی)، همون افستِ ذخیره‌شده روی شماره‌ی داخلی دیتابیس اعمال می‌شه
    offset = load_shop_settings().get("invoice_number_offset", 0) or 0
    number_value = invoice_id + offset
    invoice_number = str(number_value) if not prefix else f"{prefix}-{number_value:05d}"
    c.execute("UPDATE invoices SET number=? WHERE id=?", (invoice_number, invoice_id))

    for it in d["items"]:
        line_total = it["qty"] * it["unit_price"]
        c.execute("""INSERT INTO invoice_items (invoice_id, item_id, qty, unit_price, total, serial_number, warranty_months)
                     VALUES (?,?,?,?,?,?,?)""",
                  (invoice_id, it["item_id"], it["qty"], it["unit_price"], line_total,
                   it.get("serial_number"), it.get("warranty_months")))
        # به‌روزرسانی موجودی انبار (مرجوعی برعکسِ حالت عادی عمل می‌کند)
        sale_like = invoice_type in ("sale", "purchase_return")
        if sale_like:
            c.execute("UPDATE items SET stock_qty = stock_qty - ? WHERE id=?", (it["qty"], it["item_id"]))
        else:
            c.execute("UPDATE items SET stock_qty = stock_qty + ? WHERE id=?", (it["qty"], it["item_id"]))
            if invoice_type == "purchase":
                # میانگین موزون قیمت خرید: هر بار که خرید جدید می‌آید، با موجودی قبلی ترکیب می‌شود
                # تا هزینه‌ی واقعی‌تری برای محاسبه‌ی سود به‌دست بیاید (نه فقط آخرین قیمت خرید)
                row = c.execute("SELECT stock_qty, avg_cost, purchase_price FROM items WHERE id=?", (it["item_id"],)).fetchone()
                old_qty = (row["stock_qty"] or 0) - it["qty"]  # موجودی قبل از این خرید (چون بالا آپدیت شد)
                old_cost = row["avg_cost"] if row["avg_cost"] is not None else row["purchase_price"]
                if old_qty > 0 and old_cost is not None:
                    new_avg = (old_qty * old_cost + it["qty"] * it["unit_price"]) / (old_qty + it["qty"])
                else:
                    new_avg = it["unit_price"]
                c.execute("UPDATE items SET avg_cost=? WHERE id=?", (new_avg, it["item_id"]))

    # به‌روزرسانی حساب طرف‌حساب در صورت نسیه (برای مرجوعی، اثر معکوس اعمال می‌شود)
    if d.get("party_id") and d.get("payment_type") in ("credit",):
        remaining = total - paid
        sign = 1 if invoice_type == "sale" else (-1 if invoice_type == "purchase" else 0)
        if invoice_type == "sale_return":
            sign = -1  # بدهی مشتری کم می‌شود
        elif invoice_type == "purchase_return":
            sign = 1  # طلب ما از تامین‌کننده (کاهش بدهی ما) به‌صورت ساده در نظر گرفته می‌شود
        c.execute("UPDATE parties SET balance = balance + ? WHERE id=?", (sign * remaining, d["party_id"]))

    # ثبت در صندوق در صورت نقدی/بخشی نقد
    if paid > 0:
        tx_type = "in" if invoice_type in ("sale", "purchase_return") else "out"
        label = {"sale": "فروش", "purchase": "خرید", "sale_return": "مرجوعی فروش", "purchase_return": "مرجوعی خرید"}[invoice_type]
        c.execute("INSERT INTO cash_transactions (date, tx_type, amount, description, invoice_id) VALUES (?,?,?,?,?)",
                  (now(), tx_type, paid, f"فاکتور {label} شماره {invoice_number}", invoice_id))

    if invoice_type == "sale":
        sold_item_ids = {it["item_id"] for it in d["items"]}
        newly_out = c.execute(
            f"SELECT name, stock_qty FROM items WHERE id IN ({','.join('?' * len(sold_item_ids))}) AND stock_qty <= 0",
            tuple(sold_item_ids)
        ).fetchall()
        if newly_out:
            names = "، ".join(r["name"] for r in newly_out)
            notify_telegram_async(f"⚠️ موجودی این کالا(ها) تمام شد: {names}")

    conn.commit()
    conn.close()
    log_action(d.get("username"), f"ثبت فاکتور {invoice_number}", f"جمع کل: {total:,.0f} تومان")
    return jsonify({"ok": True, "invoice_id": invoice_id, "invoice_number": invoice_number, "total": total})


@app.route("/invoices/<int:invoice_id>", methods=["PUT"])
def update_invoice(invoice_id):
    """
    ویرایش کامل یک فاکتور موجود (اقلام، تعداد/قیمت، طرف‌حساب، نوع پرداخت، تخفیف، توضیحات).
    نوع فاکتور (فروش/خرید/مرجوعی فروش/مرجوعی خرید) قابل تغییر نیست — برای آن باید فاکتور
    حذف و دوباره با نوع درست ثبت شود.

    روش کار: ابتدا اثرات فاکتور قبلی (موجودی انبار، مانده حساب طرف‌حساب، تراکنش صندوق،
    چک‌های مرتبط) دقیقاً مثل حذف فاکتور برگردانده می‌شود؛ سپس همان اعتبارسنجی‌هایی که
    هنگام ثبت فاکتور جدید انجام می‌شود (سقف اعتبار، موجودی کافی) روی داده‌ی ویرایش‌شده
    اجرا می‌شود؛ اگر نامعتبر بود همه‌چیز rollback و خطا برگردانده می‌شود، وگرنه اثرات
    جدید اعمال و ثبت می‌شود.
    """
    err = require_admin()
    if err:
        return err
    d = request.json or {}
    conn = get_connection()
    invoice = conn.execute("SELECT * FROM invoices WHERE id=?", (invoice_id,)).fetchone()
    if not invoice:
        conn.close()
        return jsonify({"ok": False, "message": "فاکتور پیدا نشد"}), 404

    new_items = d.get("items") or []
    if not new_items:
        conn.close()
        return jsonify({"ok": False, "message": "فاکتور باید حداقل یک کالا داشته باشد"}), 400

    invoice_type = invoice["invoice_type"]  # نوع فاکتور از طریق ویرایش قابل تغییر نیست
    sale_like = invoice_type in ("sale", "purchase_return")
    old_items = conn.execute("SELECT * FROM invoice_items WHERE invoice_id=?", (invoice_id,)).fetchall()

    # ۱. برگرداندن موجودی انبار قبلی (برعکسِ همون منطقی که هنگام ثبت اعمال شده بود)
    for it in old_items:
        if sale_like:
            conn.execute("UPDATE items SET stock_qty = stock_qty + ? WHERE id=?", (it["qty"], it["item_id"]))
        else:
            conn.execute("UPDATE items SET stock_qty = stock_qty - ? WHERE id=?", (it["qty"], it["item_id"]))

    # ۲. برگرداندن مانده حساب طرف‌حساب قبلی (اگر نسیه بوده)
    if invoice["party_id"] and invoice["payment_type"] == "credit":
        old_remaining = invoice["total"] - invoice["paid"]
        sign_map = {"sale": 1, "purchase": -1, "sale_return": -1, "purchase_return": 1}
        old_sign = sign_map.get(invoice_type, 0)
        conn.execute("UPDATE parties SET balance = balance - ? WHERE id=?", (old_sign * old_remaining, invoice["party_id"]))

    # ۳. حذف تراکنش صندوق و چک‌های مرتبط قبلی (مثل حذف فاکتور)
    conn.execute("DELETE FROM cash_transactions WHERE invoice_id=?", (invoice_id,))
    related_checks = conn.execute("SELECT * FROM checks WHERE invoice_id=?", (invoice_id,)).fetchall()
    checks_note = ""
    for ch in related_checks:
        if ch["status"] == "cashed":
            conn.execute("DELETE FROM cash_transactions WHERE description=?", (f"وصول چک شماره {ch['id']}",))
            checks_note = " (چک‌های وصول‌شده مرتبط حذف و اثرشان در صندوق برگردانده شد — در صورت نیاز چک جدید ثبت کنید)"
    conn.execute("DELETE FROM checks WHERE invoice_id=?", (invoice_id,))
    conn.execute("DELETE FROM invoice_items WHERE invoice_id=?", (invoice_id,))

    # ---------- اعتبارسنجی داده‌ی جدید، روی وضعیتِ برگردانده‌شده ----------
    new_party_id = d.get("party_id")
    new_payment_type = d.get("payment_type", invoice["payment_type"])
    discount = d.get("discount", 0) or 0
    subtotal = sum(it["qty"] * it["unit_price"] for it in new_items)
    total = max(subtotal - discount, 0)
    paid = total if new_payment_type == "cash" else d.get("paid", 0)

    if invoice_type == "sale" and new_payment_type == "credit" and new_party_id:
        party = conn.execute("SELECT * FROM parties WHERE id=?", (new_party_id,)).fetchone()
        if party and party["credit_limit"] and party["credit_limit"] > 0:
            remaining = total - paid
            projected_balance = party["balance"] + remaining
            if projected_balance > party["credit_limit"]:
                conn.rollback()
                conn.close()
                return jsonify({
                    "ok": False,
                    "message": f'این فاکتور از سقف اعتبار «{party["name"]}» عبور می‌کند '
                               f'(سقف: {party["credit_limit"]:,.0f}، بدهی فعلی: {party["balance"]:,.0f}، بعد از این فاکتور: {projected_balance:,.0f})'
                }), 400

    if invoice_type == "sale":
        needed = {}
        for it in new_items:
            needed[it["item_id"]] = needed.get(it["item_id"], 0) + it["qty"]
        for item_id, qty_needed in needed.items():
            row = conn.execute("SELECT name, stock_qty FROM items WHERE id=?", (item_id,)).fetchone()
            if row and row["stock_qty"] < qty_needed:
                conn.rollback()
                conn.close()
                return jsonify({
                    "ok": False,
                    "message": f'موجودی «{row["name"]}» فقط {row["stock_qty"]:g} عدد است — نمی‌توان {qty_needed:g} تا فروخت'
                }), 400

    # ---------- اعمال اثرات جدید ----------
    for it in new_items:
        line_total = it["qty"] * it["unit_price"]
        conn.execute("""INSERT INTO invoice_items (invoice_id, item_id, qty, unit_price, total, serial_number, warranty_months)
                        VALUES (?,?,?,?,?,?,?)""",
                     (invoice_id, it["item_id"], it["qty"], it["unit_price"], line_total,
                      it.get("serial_number"), it.get("warranty_months")))
        if sale_like:
            conn.execute("UPDATE items SET stock_qty = stock_qty - ? WHERE id=?", (it["qty"], it["item_id"]))
        else:
            conn.execute("UPDATE items SET stock_qty = stock_qty + ? WHERE id=?", (it["qty"], it["item_id"]))

    if new_party_id and new_payment_type == "credit":
        remaining = total - paid
        new_sign = 1 if invoice_type == "sale" else (-1 if invoice_type == "purchase" else 0)
        if invoice_type == "sale_return":
            new_sign = -1
        elif invoice_type == "purchase_return":
            new_sign = 1
        conn.execute("UPDATE parties SET balance = balance + ? WHERE id=?", (new_sign * remaining, new_party_id))

    conn.execute(
        "UPDATE invoices SET party_id=?, total=?, paid=?, payment_type=?, description=?, discount=? WHERE id=?",
        (new_party_id, total, paid, new_payment_type, d.get("description", invoice["description"]), discount, invoice_id)
    )

    if paid > 0:
        tx_type = "in" if invoice_type in ("sale", "purchase_return") else "out"
        label = {"sale": "فروش", "purchase": "خرید", "sale_return": "مرجوعی فروش", "purchase_return": "مرجوعی خرید"}[invoice_type]
        conn.execute("INSERT INTO cash_transactions (date, tx_type, amount, description, invoice_id) VALUES (?,?,?,?,?)",
                     (now(), tx_type, paid, f"فاکتور {label} شماره {invoice['number']} (ویرایش‌شده)", invoice_id))

    conn.commit()
    conn.close()
    log_action(d.get("username"), "ویرایش فاکتور", f'فاکتور شماره {invoice["number"] or invoice_id}{checks_note}')
    log_security_event(d.get("username"), "ویرایش فاکتور",
                        f'فاکتور شماره {invoice["number"] or invoice_id} — جمع کل جدید: {total:,.0f} تومان{checks_note}')
    return jsonify({"ok": True, "invoice_id": invoice_id, "invoice_number": invoice["number"], "total": total})


@app.route("/invoices/<int:invoice_id>", methods=["DELETE"])
def delete_invoice(invoice_id):
    """
    حذف کامل یک فاکتور اشتباه: موجودی انبار، مانده حساب طرف‌حساب و تراکنش صندوق مرتبط
    همه به حالت قبل از ثبت این فاکتور برمی‌گردند.
    """
    err = require_admin()
    if err:
        return err
    d = request.json or {}
    conn = get_connection()
    invoice = conn.execute("SELECT * FROM invoices WHERE id=?", (invoice_id,)).fetchone()
    if not invoice:
        conn.close()
        return jsonify({"ok": False, "message": "فاکتور پیدا نشد"}), 404

    items = conn.execute("SELECT * FROM invoice_items WHERE invoice_id=?", (invoice_id,)).fetchall()
    invoice_type = invoice["invoice_type"]

    # ۱. برگرداندن موجودی انبار (برعکسِ همون منطقی که هنگام ثبت اعمال شده بود)
    sale_like = invoice_type in ("sale", "purchase_return")
    for it in items:
        if sale_like:
            conn.execute("UPDATE items SET stock_qty = stock_qty + ? WHERE id=?", (it["qty"], it["item_id"]))
        else:
            conn.execute("UPDATE items SET stock_qty = stock_qty - ? WHERE id=?", (it["qty"], it["item_id"]))

    # ۲. برگرداندن مانده حساب طرف‌حساب (اگر نسیه بوده)
    if invoice["party_id"] and invoice["payment_type"] == "credit":
        remaining = invoice["total"] - invoice["paid"]
        sign_map = {"sale": 1, "purchase": -1, "sale_return": -1, "purchase_return": 1}
        sign = sign_map.get(invoice_type, 0)
        conn.execute("UPDATE parties SET balance = balance - ? WHERE id=?", (sign * remaining, invoice["party_id"]))

    # ۳. حذف تراکنش صندوق مرتبط (اگر برای فاکتورهای جدیدتر ثبت شده باشد)
    conn.execute("DELETE FROM cash_transactions WHERE invoice_id=?", (invoice_id,))

    # ۳.۵ حذف چک‌های مرتبط با این فاکتور (و برگرداندن اثر صندوقی اگر قبلاً وصول شده بودند)
    related_checks = conn.execute("SELECT * FROM checks WHERE invoice_id=?", (invoice_id,)).fetchall()
    checks_note = ""
    for ch in related_checks:
        if ch["status"] == "cashed":
            conn.execute("DELETE FROM cash_transactions WHERE description=?", (f"وصول چک شماره {ch['id']}",))
            checks_note = " (یک چک وصول‌شده مرتبط هم حذف و اثرش در صندوق برگردانده شد)"
    conn.execute("DELETE FROM checks WHERE invoice_id=?", (invoice_id,))

    # ۴. حذف خود فاکتور و ردیف‌هایش
    conn.execute("DELETE FROM invoice_items WHERE invoice_id=?", (invoice_id,))
    conn.execute("DELETE FROM invoices WHERE id=?", (invoice_id,))
    conn.commit()
    conn.close()

    log_action(d.get("username"), "حذف فاکتور", f'فاکتور شماره {invoice["number"] or invoice_id}{checks_note}')
    log_security_event(d.get("username"), "حذف فاکتور", f'فاکتور شماره {invoice["number"] or invoice_id} — مبلغ {invoice["total"]:,.0f}{checks_note}')
    return jsonify({"ok": True, "message": f"فاکتور و همه موارد مرتبط حذف شد{checks_note}"})


@app.route("/invoices/<int:invoice_id>/void", methods=["POST"])
def void_invoice(invoice_id):
    """
    باطل کردن فاکتور با ثبت دلیل: برخلاف حذف کامل، خودِ فاکتور برای سابقه/حسابرسی
    باقی می‌ماند (با برچسب «باطل‌شده» و دلیل)، ولی اثرات آن (موجودی انبار، مانده حساب
    طرف‌حساب، تراکنش صندوق) دقیقاً مثل حذف برگردانده می‌شود و در گزارش‌ها لحاظ نمی‌شود.
    """
    err = require_admin()
    if err:
        return err
    d = request.json or {}
    reason = (d.get("reason") or "").strip()
    if not reason:
        return jsonify({"ok": False, "message": "برای باطل کردن فاکتور، دلیل را وارد کنید"}), 400

    conn = get_connection()
    invoice = conn.execute("SELECT * FROM invoices WHERE id=?", (invoice_id,)).fetchone()
    if not invoice:
        conn.close()
        return jsonify({"ok": False, "message": "فاکتور پیدا نشد"}), 404
    if invoice["voided"]:
        conn.close()
        return jsonify({"ok": False, "message": "این فاکتور قبلاً باطل شده است"}), 400

    items = conn.execute("SELECT * FROM invoice_items WHERE invoice_id=?", (invoice_id,)).fetchall()
    invoice_type = invoice["invoice_type"]

    sale_like = invoice_type in ("sale", "purchase_return")
    for it in items:
        if sale_like:
            conn.execute("UPDATE items SET stock_qty = stock_qty + ? WHERE id=?", (it["qty"], it["item_id"]))
        else:
            conn.execute("UPDATE items SET stock_qty = stock_qty - ? WHERE id=?", (it["qty"], it["item_id"]))

    if invoice["party_id"] and invoice["payment_type"] == "credit":
        remaining = invoice["total"] - invoice["paid"]
        sign_map = {"sale": 1, "purchase": -1, "sale_return": -1, "purchase_return": 1}
        sign = sign_map.get(invoice_type, 0)
        conn.execute("UPDATE parties SET balance = balance - ? WHERE id=?", (sign * remaining, invoice["party_id"]))

    conn.execute("DELETE FROM cash_transactions WHERE invoice_id=?", (invoice_id,))

    related_checks = conn.execute("SELECT * FROM checks WHERE invoice_id=?", (invoice_id,)).fetchall()
    checks_note = ""
    for ch in related_checks:
        if ch["status"] == "cashed":
            conn.execute("DELETE FROM cash_transactions WHERE description=?", (f"وصول چک شماره {ch['id']}",))
        checks_note = " (چک‌های مرتبط هم لغو شدند)"
    conn.execute("DELETE FROM checks WHERE invoice_id=?", (invoice_id,))

    conn.execute(
        "UPDATE invoices SET voided=1, void_reason=?, voided_by=?, voided_at=? WHERE id=?",
        (reason, g.current_user["username"], now(), invoice_id)
    )
    conn.commit()
    conn.close()

    log_action(g.current_user["username"], "باطل کردن فاکتور", f'فاکتور شماره {invoice["number"] or invoice_id} — دلیل: {reason}{checks_note}')
    log_security_event(g.current_user["username"], "باطل کردن فاکتور", f'فاکتور شماره {invoice["number"] or invoice_id} — مبلغ {invoice["total"]:,.0f} — دلیل: {reason}')
    return jsonify({"ok": True, "message": f"فاکتور باطل شد{checks_note}"})


# ---------- صندوق ----------
# ---------- بانک ----------
@app.route("/bank-accounts", methods=["GET"])
def get_bank_accounts():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM bank_accounts ORDER BY id").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/bank-accounts", methods=["POST"])
def add_bank_account():
    d = request.json
    if not d.get("name"):
        return jsonify({"ok": False, "message": "نام حساب الزامی است"}), 400
    conn = get_connection()
    conn.execute("INSERT INTO bank_accounts (name, bank_name, account_number, iban, balance, created_at) VALUES (?,?,?,?,?,?)",
                 (d["name"], d.get("bank_name", ""), d.get("account_number", ""), d.get("iban", ""), d.get("balance", 0) or 0, now()))
    conn.commit()
    conn.close()
    log_action(d.get("username"), "افزودن حساب بانکی", d["name"])
    return jsonify({"ok": True})


@app.route("/bank-accounts/<int:account_id>", methods=["PUT"])
def update_bank_account(account_id):
    d = request.json
    if not d.get("name"):
        return jsonify({"ok": False, "message": "نام حساب الزامی است"}), 400
    conn = get_connection()
    conn.execute("UPDATE bank_accounts SET name=?, bank_name=?, account_number=?, iban=? WHERE id=?",
                 (d["name"], d.get("bank_name", ""), d.get("account_number", ""), d.get("iban", ""), account_id))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/bank-accounts/<int:account_id>", methods=["DELETE"])
def delete_bank_account(account_id):
    conn = get_connection()
    acc = conn.execute("SELECT * FROM bank_accounts WHERE id=?", (account_id,)).fetchone()
    if acc and acc["balance"] != 0:
        conn.close()
        return jsonify({"ok": False, "message": "فقط حساب با موجودی صفر قابل حذف است"}), 400
    conn.execute("DELETE FROM bank_accounts WHERE id=?", (account_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


BANK_DEPOSIT_SOURCE_LABELS = {
    "customer": "دریافت از مشتری",
    "person_transfer": "واریزی/انتقال از شخص یا حساب دیگر",
    "capital": "آورده/سرمایه‌گذاری صاحب کسب‌وکار",
    "interest": "سود بانکی",
    "other": "سایر",
}
BANK_WITHDRAWAL_DEST_LABELS = {
    "person_transfer": "انتقال به حساب شخص یا شرکت دیگر",
    "rent": "اجاره",
    "salary": "حقوق پرسنل",
    "utilities": "قبوض",
    "repairs": "تعمیر و نگهداری",
    "transport": "حمل و نقل",
    "supplies": "لوازم مصرفی مغازه",
    "bank_fee": "کارمزد بانکی",
    "personal_draw": "برداشت شخصی/سود مالک",
    "other": "متفرقه",
}


@app.route("/bank-accounts/<int:account_id>/transaction", methods=["POST"])
def bank_account_transaction(account_id):
    """واریز یا برداشت مستقیم از یک حساب بانکی (بدون ارتباط با صندوق)"""
    d = request.json
    tx_type = d.get("tx_type")  # deposit یا withdrawal
    amount = d.get("amount", 0)
    if tx_type not in ("deposit", "withdrawal") or amount <= 0:
        return jsonify({"ok": False, "message": "اطلاعات تراکنش نامعتبر است"}), 400

    conn = get_connection()
    acc = conn.execute("SELECT * FROM bank_accounts WHERE id=?", (account_id,)).fetchone()
    if not acc:
        conn.close()
        return jsonify({"ok": False, "message": "حساب پیدا نشد"}), 404
    if tx_type == "withdrawal" and acc["balance"] < amount:
        conn.close()
        return jsonify({"ok": False, "message": f'موجودی حساب کافی نیست (موجودی فعلی: {acc["balance"]:,.0f})'}), 400

    category = d.get("category") or None
    labels = BANK_DEPOSIT_SOURCE_LABELS if tx_type == "deposit" else BANK_WITHDRAWAL_DEST_LABELS
    category_label = labels.get(category)
    user_desc = (d.get("description") or "").strip()
    if category_label:
        desc = category_label + (f" — {user_desc}" if user_desc else "")
    else:
        desc = user_desc

    sign = 1 if tx_type == "deposit" else -1
    conn.execute("UPDATE bank_accounts SET balance = balance + ? WHERE id=?", (sign * amount, account_id))
    conn.execute("INSERT INTO bank_transactions (account_id, date, tx_type, amount, description, username, category) VALUES (?,?,?,?,?,?,?)",
                 (account_id, now(), tx_type, amount, desc, d.get("username"), category))
    conn.commit()
    conn.close()
    log_action(d.get("username"), "تراکنش بانکی",
               f'{"واریز" if tx_type == "deposit" else "برداشت"} {amount:,.0f} تومان — {acc["name"]}'
               + (f' ({category_label})' if category_label else ''))
    return jsonify({"ok": True})


@app.route("/bank-accounts/<int:account_id>/statement", methods=["GET"])
def bank_account_statement(account_id):
    conn = get_connection()
    acc = conn.execute("SELECT * FROM bank_accounts WHERE id=?", (account_id,)).fetchone()
    if not acc:
        conn.close()
        return jsonify({"ok": False, "message": "حساب پیدا نشد"}), 404
    rows = conn.execute("SELECT * FROM bank_transactions WHERE account_id=? ORDER BY date DESC, id DESC", (account_id,)).fetchall()
    conn.close()
    return jsonify({"account": dict(acc), "transactions": [dict(r) for r in rows]})


@app.route("/bank-transfer", methods=["POST"])
def bank_transfer():
    """
    انتقال بین صندوق و بانک یا بین دو حساب بانکی.
    ورودی: {"from_type": "cash"|"bank", "from_id": null یا شماره حساب,
             "to_type": "cash"|"bank", "to_id": null یا شماره حساب, "amount": عدد, "description": "..."}
    """
    d = request.json
    amount = d.get("amount", 0)
    from_type, to_type = d.get("from_type"), d.get("to_type")
    if amount <= 0 or from_type == to_type == "cash":
        return jsonify({"ok": False, "message": "اطلاعات انتقال نامعتبر است"}), 400

    conn = get_connection()

    # بررسی موجودی کافی مبدأ
    if from_type == "cash":
        bal = conn.execute("SELECT COALESCE(SUM(CASE WHEN tx_type='in' THEN amount ELSE -amount END),0) as b FROM cash_transactions").fetchone()["b"]
        if bal < amount:
            conn.close()
            return jsonify({"ok": False, "message": f"موجودی صندوق کافی نیست (موجودی فعلی: {bal:,.0f})"}), 400
    else:
        acc = conn.execute("SELECT * FROM bank_accounts WHERE id=?", (d.get("from_id"),)).fetchone()
        if not acc or acc["balance"] < amount:
            conn.close()
            return jsonify({"ok": False, "message": "موجودی حساب بانکی مبدأ کافی نیست"}), 400

    def account_label(acc_type, acc_id):
        if acc_type == "cash":
            return "صندوق"
        row = conn.execute("SELECT name FROM bank_accounts WHERE id=?", (acc_id,)).fetchone()
        return row["name"] if row else "حساب بانکی"

    from_label = account_label(from_type, d.get("from_id"))
    to_label = account_label(to_type, d.get("to_id"))
    user_desc = d.get("description")
    desc_out = f"انتقال به «{to_label}»" + (f" — {user_desc}" if user_desc else "")
    desc_in = f"انتقال از «{from_label}»" + (f" — {user_desc}" if user_desc else "")

    # کسر از مبدأ
    if from_type == "cash":
        conn.execute("INSERT INTO cash_transactions (date, tx_type, amount, description) VALUES (?,?,?,?)",
                     (now(), "out", amount, desc_out))
    else:
        conn.execute("UPDATE bank_accounts SET balance = balance - ? WHERE id=?", (amount, d.get("from_id")))
        conn.execute("INSERT INTO bank_transactions (account_id, date, tx_type, amount, description, username) VALUES (?,?,?,?,?,?)",
                     (d.get("from_id"), now(), "transfer_out", amount, desc_out, d.get("username")))

    # افزودن به مقصد
    if to_type == "cash":
        conn.execute("INSERT INTO cash_transactions (date, tx_type, amount, description) VALUES (?,?,?,?)",
                     (now(), "in", amount, desc_in))
    else:
        conn.execute("UPDATE bank_accounts SET balance = balance + ? WHERE id=?", (amount, d.get("to_id")))
        conn.execute("INSERT INTO bank_transactions (account_id, date, tx_type, amount, description, username) VALUES (?,?,?,?,?,?)",
                     (d.get("to_id"), now(), "transfer_in", amount, desc_in, d.get("username")))

    conn.commit()
    conn.close()
    log_action(d.get("username"), "انتقال وجه", f'{amount:,.0f} تومان — از «{from_label}» به «{to_label}»')
    return jsonify({"ok": True})


@app.route("/cash", methods=["GET"])
def get_cash():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM cash_transactions ORDER BY date DESC, id DESC").fetchall()
    balance = conn.execute("""SELECT
        COALESCE(SUM(CASE WHEN tx_type='in' THEN amount ELSE -amount END),0) as bal
        FROM cash_transactions""").fetchone()["bal"]
    conn.close()
    return jsonify({"transactions": [dict(r) for r in rows], "balance": balance})


@app.route("/cash", methods=["POST"])
def add_cash():
    err = require_permission("can_manage_cash")
    if err:
        return err
    d = request.json
    conn = get_connection()
    # دسته‌بندی هزینه فقط برای تراکنش‌های خروج دستی معنا دارد (نه پرداخت‌های مربوط به فاکتور
    # که خودکار ثبت می‌شوند)؛ همین باعث می‌شه محاسبه‌ی سود، هزینه‌ها رو دوبار حساب نکنه
    expense_category = d.get("expense_category") if d.get("tx_type") == "out" else None
    conn.execute("INSERT INTO cash_transactions (date, tx_type, amount, description, expense_category) VALUES (?,?,?,?,?)",
                 (now(), d["tx_type"], d["amount"], d.get("description", ""), expense_category))
    conn.commit()
    conn.close()
    log_action(d.get("username"), "ثبت تراکنش صندوق",
               f'{"دریافت" if d["tx_type"]=="in" else "پرداخت"}: {d["amount"]:,.0f} تومان'
               + (f' ({expense_category})' if expense_category else ''))
    return jsonify({"ok": True})


@app.route("/cash/closings", methods=["GET"])
def get_cash_closings():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM cash_closings ORDER BY id DESC LIMIT 90").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/cash/closings", methods=["POST"])
def add_cash_closing():
    """
    بستن روزانه صندوق: مبلغ نقد واقعی شمرده‌شده در کشو رو با موجودی محاسبه‌شده توسط
    سیستم مقایسه می‌کنه و اختلاف (اگه بود) رو ثبت می‌کنه — برای کشف زودهنگام کم‌فروشی
    یا اشتباه در پول برگشت.
    """
    d = request.json or {}
    try:
        counted = float(d.get("counted_balance"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "message": "مبلغ شمرده‌شده باید عدد باشد"}), 400

    conn = get_connection()
    expected = conn.execute("""SELECT
        COALESCE(SUM(CASE WHEN tx_type='in' THEN amount ELSE -amount END),0) as bal
        FROM cash_transactions""").fetchone()["bal"]
    difference = counted - expected
    conn.execute(
        "INSERT INTO cash_closings (date, expected_balance, counted_balance, difference, note, username, created_at) "
        "VALUES (?,?,?,?,?,?,?)",
        (d.get("date") or now()[:10], expected, counted, difference, d.get("note", ""), d.get("username"), now())
    )
    conn.commit()
    conn.close()
    log_action(d.get("username"), "بستن روزانه صندوق",
               f'محاسبه‌شده: {expected:,.0f} — شمرده‌شده: {counted:,.0f} — اختلاف: {difference:,.0f} تومان')
    if abs(difference) > 0:
        log_security_event(d.get("username"), "اختلاف در بستن صندوق",
                            f'اختلاف {difference:,.0f} تومان (محاسبه‌شده: {expected:,.0f}، شمرده‌شده: {counted:,.0f})')
    return jsonify({"ok": True, "expected_balance": expected, "difference": difference})


@app.route("/invoices/<int:invoice_id>/installments", methods=["GET"])
def get_installments(invoice_id):
    conn = get_connection()
    rows = conn.execute("SELECT * FROM invoice_installments WHERE invoice_id=? ORDER BY due_date", (invoice_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/invoices/<int:invoice_id>/installments", methods=["POST"])
def set_installments(invoice_id):
    """
    تعریف اقساط برای فاکتور نسیه: مانده‌ی فاکتور را به چند قسط با تاریخ سررسید تقسیم می‌کند.
    اقساط قبلی این فاکتور (اگر بود) جایگزین می‌شوند.
    """
    err = require_admin()
    if err:
        return err
    d = request.json or {}
    installments = d.get("installments") or []
    if not installments:
        return jsonify({"ok": False, "message": "حداقل یک قسط لازم است"}), 400
    conn = get_connection()
    invoice = conn.execute("SELECT * FROM invoices WHERE id=?", (invoice_id,)).fetchone()
    if not invoice:
        conn.close()
        return jsonify({"ok": False, "message": "فاکتور پیدا نشد"}), 404
    if invoice["invoice_type"] != "sale" or invoice["payment_type"] != "credit":
        conn.close()
        return jsonify({"ok": False, "message": "اقساط فقط برای فاکتور فروش نسیه قابل تعریف است"}), 400
    remaining = invoice["total"] - invoice["paid"]
    total_installments = sum(i["amount"] for i in installments)
    if abs(total_installments - remaining) > 1:
        conn.close()
        return jsonify({"ok": False, "message": f"جمع اقساط ({total_installments:,.0f}) باید برابر مانده فاکتور ({remaining:,.0f}) باشد"}), 400
    conn.execute("DELETE FROM invoice_installments WHERE invoice_id=?", (invoice_id,))
    for i in installments:
        conn.execute("INSERT INTO invoice_installments (invoice_id, due_date, amount) VALUES (?,?,?)",
                     (invoice_id, i["due_date"], i["amount"]))
    conn.commit()
    conn.close()
    log_action(d.get("username"), "تعریف اقساط فاکتور", f'فاکتور شماره {invoice["number"] or invoice_id} — {len(installments)} قسط')
    return jsonify({"ok": True})


@app.route("/installments/<int:installment_id>/pay", methods=["POST"])
def pay_installment(installment_id):
    """پرداخت یک قسط: مبلغش به صندوق واریز و از بدهی طرف‌حساب کم می‌شود"""
    d = request.json or {}
    conn = get_connection()
    inst = conn.execute("SELECT * FROM invoice_installments WHERE id=?", (installment_id,)).fetchone()
    if not inst:
        conn.close()
        return jsonify({"ok": False, "message": "قسط پیدا نشد"}), 404
    if inst["paid"]:
        conn.close()
        return jsonify({"ok": False, "message": "این قسط قبلاً پرداخت شده است"}), 400
    invoice = conn.execute("SELECT * FROM invoices WHERE id=?", (inst["invoice_id"],)).fetchone()
    conn.execute("UPDATE invoice_installments SET paid=1, paid_at=? WHERE id=?", (now(), installment_id))
    conn.execute("INSERT INTO cash_transactions (date, tx_type, amount, description, invoice_id) VALUES (?,?,?,?,?)",
                 (now(), "in", inst["amount"], f'پرداخت قسط فاکتور شماره {invoice["number"] or invoice["id"]}', invoice["id"]))
    if invoice["party_id"]:
        conn.execute("UPDATE parties SET balance = balance - ? WHERE id=?", (inst["amount"], invoice["party_id"]))
    conn.execute("UPDATE invoices SET paid = paid + ? WHERE id=?", (inst["amount"], invoice["id"]))
    conn.commit()
    conn.close()
    log_action(d.get("username"), "پرداخت قسط", f'فاکتور شماره {invoice["number"] or invoice["id"]} — مبلغ {inst["amount"]:,.0f} تومان')
    return jsonify({"ok": True})


@app.route("/items/stocktake", methods=["POST"])
def apply_stocktake():
    """
    انبارگردانی: تعدادهای شمارش‌شده (با اسکن بارکد) را با موجودی فعلی سیستم مقایسه
    و موجودی را با تعداد واقعی شمرده‌شده جایگزین می‌کند (نه جمع می‌زند).
    """
    err = require_admin()
    if err:
        return err
    d = request.json or {}
    counts = d.get("counts") or {}
    if not counts:
        return jsonify({"ok": False, "message": "لیست شمارش خالی است"}), 400
    conn = get_connection()
    changes = []
    for item_id_str, counted_qty in counts.items():
        item_id = int(item_id_str)
        item = conn.execute("SELECT name, stock_qty FROM items WHERE id=?", (item_id,)).fetchone()
        if not item:
            continue
        old_qty = item["stock_qty"]
        if old_qty != counted_qty:
            conn.execute("UPDATE items SET stock_qty=? WHERE id=?", (counted_qty, item_id))
            changes.append(f'{item["name"]}: {old_qty:g} → {counted_qty:g}')
    conn.commit()
    conn.close()
    if changes:
        log_action(d.get("username"), "انبارگردانی", "؛ ".join(changes))
        log_security_event(d.get("username"), "اصلاح موجودی (انبارگردانی)", "؛ ".join(changes))
    return jsonify({"ok": True, "changed_count": len(changes)})


@app.route("/warranty-claims", methods=["GET"])
def get_warranty_claims():
    conn = get_connection()
    rows = conn.execute("""
        SELECT warranty_claims.*, items.name as item_name, parties.name as party_name
        FROM warranty_claims
        LEFT JOIN items ON warranty_claims.item_id = items.id
        LEFT JOIN parties ON warranty_claims.party_id = parties.id
        ORDER BY warranty_claims.id DESC
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/warranty-claims", methods=["POST"])
def add_warranty_claim():
    d = request.json or {}
    if not (d.get("item_id") or d.get("serial_number")):
        return jsonify({"ok": False, "message": "کالا یا شماره سریال را مشخص کنید"}), 400
    conn = get_connection()
    conn.execute("""
        INSERT INTO warranty_claims (item_id, serial_number, party_id, invoice_id, issue_description, status, created_at, username)
        VALUES (?,?,?,?,?,?,?,?)
    """, (d.get("item_id"), d.get("serial_number"), d.get("party_id"), d.get("invoice_id"),
          d.get("issue_description", ""), "received", now(), d.get("username")))
    conn.commit()
    conn.close()
    log_action(d.get("username"), "ثبت درخواست گارانتی/تعمیر", d.get("issue_description", ""))
    return jsonify({"ok": True})


@app.route("/warranty-claims/<int:claim_id>", methods=["PUT"])
def update_warranty_claim(claim_id):
    d = request.json or {}
    conn = get_connection()
    claim = conn.execute("SELECT * FROM warranty_claims WHERE id=?", (claim_id,)).fetchone()
    if not claim:
        conn.close()
        return jsonify({"ok": False, "message": "درخواست پیدا نشد"}), 404
    status = d.get("status", claim["status"])
    resolved_at = now() if status in ("done", "returned") and claim["status"] not in ("done", "returned") else claim["resolved_at"]
    conn.execute(
        "UPDATE warranty_claims SET status=?, note=?, resolved_at=? WHERE id=?",
        (status, d.get("note", claim["note"]), resolved_at, claim_id)
    )
    conn.commit()
    conn.close()
    log_action(d.get("username"), "به‌روزرسانی درخواست گارانتی/تعمیر", f"وضعیت: {status}")
    return jsonify({"ok": True})


@app.route("/reports/expenses", methods=["GET"])
def report_expenses():
    """جمع هزینه‌های جاری (خروج دستی صندوق با دسته‌بندی) به تفکیک دسته، در N روز اخیر"""
    err = require_admin()
    if err:
        return err
    days = request.args.get("days", default=30, type=int)
    conn = get_connection()
    since = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    rows = conn.execute("""
        SELECT COALESCE(expense_category, 'سایر') as category, SUM(amount) as total_amount, COUNT(*) as tx_count
        FROM cash_transactions
        WHERE tx_type='out' AND invoice_id IS NULL AND expense_category IS NOT NULL AND date >= ?
        GROUP BY category
        ORDER BY total_amount DESC
    """, (since,)).fetchall()
    total = conn.execute("""
        SELECT COALESCE(SUM(amount),0) as t FROM cash_transactions
        WHERE tx_type='out' AND invoice_id IS NULL AND expense_category IS NOT NULL AND date >= ?
    """, (since,)).fetchone()["t"]
    conn.close()
    return jsonify({"categories": [dict(r) for r in rows], "total": total})


# ---------- گزارش‌ها ----------
@app.route("/reports/summary", methods=["GET"])
def report_summary():
    conn = get_connection()
    sales = conn.execute("SELECT COALESCE(SUM(total),0) as t FROM invoices WHERE invoice_type='sale' AND voided=0").fetchone()["t"]
    sales_returns = conn.execute("SELECT COALESCE(SUM(total),0) as t FROM invoices WHERE invoice_type='sale_return' AND voided=0").fetchone()["t"]
    purchases = conn.execute("SELECT COALESCE(SUM(total),0) as t FROM invoices WHERE invoice_type='purchase' AND voided=0").fetchone()["t"]
    purchase_returns = conn.execute("SELECT COALESCE(SUM(total),0) as t FROM invoices WHERE invoice_type='purchase_return' AND voided=0").fetchone()["t"]
    net_sales = sales - sales_returns
    net_purchases = purchases - purchase_returns
    operating_expenses = conn.execute("""
        SELECT COALESCE(SUM(amount),0) as t FROM cash_transactions
        WHERE tx_type='out' AND invoice_id IS NULL AND expense_category IS NOT NULL
    """).fetchone()["t"]
    debtors = conn.execute("SELECT COALESCE(SUM(balance),0) as t FROM parties WHERE balance > 0").fetchone()["t"]
    low_stock = conn.execute("SELECT * FROM items WHERE stock_qty <= min_stock AND deleted_at IS NULL").fetchall()
    conn.close()
    is_admin = g.current_user.get("role") == "admin"
    return jsonify({
        "total_sales": net_sales,
        "total_purchases": net_purchases if is_admin else None,
        "total_operating_expenses": operating_expenses if is_admin else None,
        "estimated_profit": (net_sales - net_purchases - operating_expenses) if is_admin else None,
        "total_debtors": debtors,
        "low_stock_items": [dict(r) for r in low_stock]
    })


@app.route("/reports/stock-ledger/<int:item_id>", methods=["GET"])
def stock_ledger(item_id):
    """گردش انبار یک کالا: همه ورود/خروج‌های آن به ترتیب تاریخ"""
    conn = get_connection()
    item = conn.execute("SELECT * FROM items WHERE id=?", (item_id,)).fetchone()
    if not item:
        conn.close()
        return jsonify({"ok": False, "message": "کالا پیدا نشد"}), 404
    rows = conn.execute("""
        SELECT invoices.date as date, invoices.invoice_type as invoice_type,
               invoices.number as number, invoice_items.qty as qty
        FROM invoice_items
        JOIN invoices ON invoice_items.invoice_id = invoices.id
        WHERE invoice_items.item_id = ? AND invoices.voided=0
        ORDER BY invoices.date
    """, (item_id,)).fetchall()
    conn.close()

    movement_label = {"sale": "خروج (فروش)", "purchase": "ورود (خرید)",
                       "sale_return": "ورود (مرجوعی فروش)", "purchase_return": "خروج (مرجوعی خرید)"}
    running = 0
    ledger = []
    for r in rows:
        qty_in = r["qty"] if r["invoice_type"] in ("purchase", "sale_return") else -r["qty"]
        running += qty_in
        ledger.append({
            "date": r["date"], "type": movement_label.get(r["invoice_type"], r["invoice_type"]),
            "number": r["number"], "qty": r["qty"], "running_balance": running
        })
    return jsonify({"item": dict(item), "ledger": ledger})


@app.route("/activity-log", methods=["GET"])
def get_activity_log():
    conn = get_connection()
    rows = conn.execute("SELECT * FROM activity_log ORDER BY id DESC LIMIT 500").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/security-log", methods=["GET"])
def get_security_log():
    err = require_admin()
    if err:
        return err
    conn = get_connection()
    rows = conn.execute("SELECT * FROM security_log ORDER BY id DESC LIMIT 500").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


# ---------- چک‌ها ----------
@app.route("/checks", methods=["GET"])
def get_checks():
    status = request.args.get("status")
    conn = get_connection()
    q = """SELECT checks.*, parties.name as party_name FROM checks
           LEFT JOIN parties ON checks.party_id = parties.id"""
    params = ()
    if status:
        q += " WHERE status=?"
        params = (status,)
    q += " ORDER BY due_date ASC"
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/checks", methods=["POST"])
def add_check():
    d = request.json
    if not d.get("amount") or not d.get("due_date") or d.get("direction") not in ("received", "issued"):
        return jsonify({"ok": False, "message": "مبلغ، تاریخ سررسید و نوع چک الزامی است"}), 400
    conn = get_connection()
    conn.execute("""INSERT INTO checks (party_id, invoice_id, amount, due_date, status, direction, description)
                     VALUES (?,?,?,?,?,?,?)""",
                 (d.get("party_id"), d.get("invoice_id"), d["amount"], d["due_date"],
                  d.get("status", "pending"), d["direction"], d.get("description", "")))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/checks/<int:check_id>", methods=["PUT"])
def update_check(check_id):
    """ویرایش چک (طرف حساب/مبلغ/سررسید/نوع/توضیحات) و/یا تغییر وضعیت (pending -> cashed یا bounced).
    هر فیلدی که در بدنه‌ی درخواست نیاید، مقدار قبلی‌اش حفظ می‌شود — پس همین مسیر هم برای
    تغییر سریع وضعیت (فقط با فرستادن status) و هم برای ویرایش کامل کاربرد دارد."""
    d = request.json
    conn = get_connection()
    check = conn.execute("SELECT * FROM checks WHERE id=?", (check_id,)).fetchone()
    if not check:
        conn.close()
        return jsonify({"ok": False, "message": "چک پیدا نشد"}), 404

    new_status = d.get("status", check["status"])
    conn.execute(
        "UPDATE checks SET party_id=?, amount=?, due_date=?, direction=?, description=?, status=? WHERE id=?",
        (d.get("party_id", check["party_id"]), d.get("amount", check["amount"]),
         d.get("due_date", check["due_date"]), d.get("direction", check["direction"]),
         d.get("description", check["description"]), new_status, check_id),
    )

    if new_status == "cashed" and check["status"] != "cashed":
        tx_type = "in" if check["direction"] == "received" else "out"
        conn.execute("INSERT INTO cash_transactions (date, tx_type, amount, description) VALUES (?,?,?,?)",
                      (now(), tx_type, check["amount"], f"وصول چک شماره {check_id}"))
    if new_status == "bounced" and check["status"] != "bounced":
        party = conn.execute("SELECT name FROM parties WHERE id=?", (check["party_id"],)).fetchone() if check["party_id"] else None
        notify_telegram_async(f"🔴 چک {check['amount']:,.0f} تومانی {party['name'] if party else ''} برگشت خورد")
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/checks/<int:check_id>", methods=["DELETE"])
def delete_check(check_id):
    conn = get_connection()
    conn.execute("DELETE FROM checks WHERE id=?", (check_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ---------- پشتیبان‌گیری ----------
def do_backup():
    if not os.path.exists(DB_PATH):
        return None
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    dest = os.path.join(BACKUP_DIR, f"accounting_{stamp}.db")
    shutil.copy2(DB_PATH, dest)
    return dest


def backup_loop():
    last_backup_date = None
    while True:
        today = datetime.now().strftime("%Y-%m-%d")
        if today != last_backup_date:
            do_backup()
            last_backup_date = today
        time.sleep(3600)


@app.route("/backup/now", methods=["POST"])
def backup_now():
    err = require_admin()
    if err:
        return err
    path = do_backup()
    return jsonify({"ok": True, "path": path})


@app.route("/backup/list", methods=["GET"])
def backup_list():
    err = require_admin()
    if err:
        return err
    files = sorted(os.listdir(BACKUP_DIR), reverse=True)
    return jsonify(files)


# ---------- گزارش شبانه (ایمیل + آپلود به VPS) ----------
def run_nightly_tasks():
    backup_path = do_backup()
    summary_text = notifier.build_daily_summary(get_connection)
    ok_email, msg_email = notifier.send_email_summary(summary_text, backup_path)
    ok_vps, msg_vps = notifier.upload_to_vps(backup_path)
    print("گزارش شبانه ->", "ایمیل:", msg_email, "| VPS:", msg_vps)
    return {"email": {"ok": ok_email, "message": msg_email},
            "vps": {"ok": ok_vps, "message": msg_vps}}


def nightly_loop():
    """هر روز ساعت ۲۳ (۱۱ شب) یک‌بار گزارش فروش را ایمیل و بکاپ را به VPS آپلود می‌کند"""
    last_run_date = None
    while True:
        now_dt = datetime.now()
        today_str = now_dt.strftime("%Y-%m-%d")
        if now_dt.hour == 23 and today_str != last_run_date:
            try:
                run_nightly_tasks()
            except Exception as e:
                print("خطا در اجرای گزارش شبانه:", e)
            last_run_date = today_str
        time.sleep(300)


@app.route("/nightly/run-now", methods=["POST"])
def nightly_run_now():
    """برای تست دستی: بلافاصله گزارش شبانه را اجرا می‌کند بدون نیاز به صبر تا ساعت ۲۳"""
    err = require_admin()
    if err:
        return err
    result = run_nightly_tasks()
    return jsonify(result)


def run_weekly_recap():
    text = notifier.build_weekly_summary(get_connection)
    notify_telegram_async(text)


def weekly_recap_loop():
    """هر یکشنبه ساعت ۹ صبح یک‌بار خلاصه‌ی هفتگی فروش را به تلگرام صاحب مغازه می‌فرستد
    (نیاز به تنظیم تلگرام در بخش تنظیمات کلی دارد؛ در غیر این صورت بی‌صدا نادیده گرفته می‌شود)"""
    last_run_date = None
    while True:
        now_dt = datetime.now()
        today_str = now_dt.strftime("%Y-%m-%d")
        if now_dt.weekday() == 6 and now_dt.hour == 9 and today_str != last_run_date:
            try:
                run_weekly_recap()
            except Exception as e:
                print("خطا در ارسال خلاصه هفتگی:", e)
            last_run_date = today_str
        time.sleep(300)


@app.route("/weekly-recap/run-now", methods=["POST"])
def weekly_recap_run_now():
    """برای تست دستی: بلافاصله خلاصه هفتگی را به تلگرام می‌فرستد"""
    err = require_admin()
    if err:
        return err
    run_weekly_recap()
    return jsonify({"ok": True})


@app.route("/reports/monthly", methods=["GET"])
def report_monthly():
    """خلاصه فروش، خرید و سود هر ماه برای ۱۲ ماه اخیر (بر اساس ماه‌های میلادی ثبت‌شده در دیتابیس)"""
    conn = get_connection()
    rows = conn.execute("""
        SELECT substr(date, 1, 7) as month,
               SUM(CASE WHEN invoice_type='sale' THEN total ELSE 0 END) as sales,
               SUM(CASE WHEN invoice_type='purchase' THEN total ELSE 0 END) as purchases
        FROM invoices
        GROUP BY month
        ORDER BY month
    """).fetchall()
    conn.close()
    is_admin = g.current_user.get("role") == "admin"
    result = []
    for r in rows:
        result.append({
            "month": r["month"],
            "sales": r["sales"] or 0,
            "purchases": (r["purchases"] or 0) if is_admin else None,
            "profit": ((r["sales"] or 0) - (r["purchases"] or 0)) if is_admin else None
        })
    return jsonify(result)


# ---------- مدیریت کاربران (برای سطح دسترسی) ----------
@app.route("/users", methods=["GET"])
def get_users():
    err = require_admin()
    if err:
        return err
    conn = get_connection()
    rows = conn.execute("SELECT id, username, role, permissions, totp_enabled FROM users ORDER BY id").fetchall()
    conn.close()
    result = []
    for r in rows:
        row = dict(r)
        try:
            row["permissions"] = json.loads(row["permissions"]) if row["permissions"] else {}
        except (TypeError, ValueError):
            row["permissions"] = {}
        result.append(row)
    return jsonify(result)


@app.route("/users/<int:user_id>/permissions", methods=["PUT"])
def update_user_permissions(user_id):
    err = require_admin()
    if err:
        return err
    d = request.json or {}
    perms = {k: bool(d.get(k, True)) for k in PERMISSION_KEYS}
    conn = get_connection()
    conn.execute("UPDATE users SET permissions=? WHERE id=?", (json.dumps(perms), user_id))
    conn.commit()
    conn.close()
    log_security_event(d.get("username"), "تغییر سطح دسترسی کاربر", f"کاربر شماره {user_id}: {perms}")
    return jsonify({"ok": True})


@app.route("/users", methods=["POST"])
def add_user():
    err = require_admin()
    if err:
        return err
    d = request.json
    conn = get_connection()
    try:
        conn.execute("INSERT INTO users (username, password, role) VALUES (?,?,?)",
                     (d["username"], generate_password_hash(d["password"]), d.get("role", "employee")))
        conn.commit()
        ok = True
        msg = "کاربر اضافه شد"
        log_security_event(d.get("username"), "افزودن کاربر جدید", f'کاربر: {d["username"]}')
    except Exception as e:
        ok = False
        msg = "این نام کاربری قبلاً استفاده شده" if "UNIQUE" in str(e) else str(e)
    conn.close()
    return jsonify({"ok": ok, "message": msg})


@app.route("/users/<int:user_id>", methods=["PUT"])
def update_user(user_id):
    err = require_admin()
    if err:
        return err
    d = request.json
    conn = get_connection()
    if d.get("password"):
        # وقتی مدیر رمز یکی دیگر رو عوض می‌کنه، اون رمز یه رمز موقته — کاربر باید بعد از
        # اولین ورود خودش یه رمز تازه انتخاب کنه
        conn.execute("UPDATE users SET password=?, role=?, must_change_password=1 WHERE id=?",
                     (generate_password_hash(d["password"]), d.get("role", "employee"), user_id))
        log_security_event(d.get("username"), "تغییر رمز عبور", f"کاربر شماره {user_id}")
    else:
        conn.execute("UPDATE users SET role=? WHERE id=?", (d.get("role", "employee"), user_id))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/users/me/password", methods=["POST"])
def change_own_password():
    """
    تغییر رمز عبور خودِ کاربر لاگین‌شده (نه ادمین‌محدود، چون هر کاربری — حتی کارمند —
    باید بتونه رمز خودش رو عوض کنه، مثلاً وقتی بعد از ریست شدن رمزش توسط مدیر مجبور
    به انتخاب رمز جدید می‌شه).
    """
    d = request.json or {}
    new_password = d.get("new_password") or ""
    if len(new_password) < 4:
        return jsonify({"ok": False, "message": "رمز عبور جدید باید حداقل ۴ کاراکتر باشد"}), 400
    username = g.current_user["username"]
    conn = get_connection()
    conn.execute("UPDATE users SET password=?, must_change_password=0 WHERE username=?",
                 (generate_password_hash(new_password), username))
    conn.commit()
    conn.close()
    log_security_event(username, "تغییر رمز عبور توسط خود کاربر", "")
    return jsonify({"ok": True})


@app.route("/users/me/2fa/setup", methods=["POST"])
def setup_2fa():
    """مرحله اول فعال‌سازی ورود دو مرحله‌ای: یک کلید تصادفی جدید می‌سازد (هنوز فعال نیست
    تا با /verify تایید شود) و کلید را برای وارد کردن دستی در اپ‌های Google Authenticator/Authy برمی‌گرداند."""
    username = g.current_user["username"]
    secret = generate_totp_secret()
    conn = get_connection()
    conn.execute("UPDATE users SET totp_secret=?, totp_enabled=0 WHERE username=?", (secret, username))
    conn.commit()
    conn.close()
    otpauth_uri = f"otpauth://totp/AccountingApp:{username}?secret={secret}&issuer=AccountingApp"
    return jsonify({"ok": True, "secret": secret, "otpauth_uri": otpauth_uri})


@app.route("/users/me/2fa/verify", methods=["POST"])
def verify_2fa():
    """مرحله دوم: کاربر یک کد از اپ احراز هویتش وارد می‌کند تا مطمئن شویم درست تنظیم شده، سپس فعال می‌شود"""
    d = request.json or {}
    username = g.current_user["username"]
    conn = get_connection()
    user = conn.execute("SELECT totp_secret FROM users WHERE username=?", (username,)).fetchone()
    if not user or not user["totp_secret"]:
        conn.close()
        return jsonify({"ok": False, "message": "ابتدا مرحله راه‌اندازی را انجام دهید"}), 400
    if not verify_totp(user["totp_secret"], d.get("code")):
        conn.close()
        return jsonify({"ok": False, "message": "کد اشتباه است"}), 400
    conn.execute("UPDATE users SET totp_enabled=1 WHERE username=?", (username,))
    conn.commit()
    conn.close()
    log_security_event(username, "فعال‌سازی ورود دو مرحله‌ای", "")
    return jsonify({"ok": True})


@app.route("/users/me/2fa/disable", methods=["POST"])
def disable_2fa():
    d = request.json or {}
    username = g.current_user["username"]
    conn = get_connection()
    user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    if not user or not check_password_hash(user["password"], d.get("password", "")):
        conn.close()
        return jsonify({"ok": False, "message": "رمز عبور اشتباه است"}), 401
    conn.execute("UPDATE users SET totp_enabled=0, totp_secret=NULL WHERE username=?", (username,))
    conn.commit()
    conn.close()
    log_security_event(username, "غیرفعال‌سازی ورود دو مرحله‌ای", "")
    return jsonify({"ok": True})


@app.route("/users/<int:user_id>", methods=["DELETE"])
def delete_user(user_id):
    err = require_admin()
    if err:
        return err
    conn = get_connection()
    conn.execute("DELETE FROM users WHERE id=?", (user_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/backup/restore", methods=["POST"])
def backup_restore():
    """بازیابی دیتابیس از یکی از نسخه‌های پشتیبان قبلی. قبل از بازیابی، خودش هم یک بکاپ احتیاطی از وضعیت فعلی می‌گیرد."""
    err = require_admin()
    if err:
        return err
    d = request.json
    filename = d.get("filename")
    if not filename:
        return jsonify({"ok": False, "message": "نام فایل پشتیبان ارسال نشده"}), 400
    # فقط نام فایل خالی (بدون مسیر) مجاز است تا کسی نتواند با ../ به فایل‌های
    # خارج از پوشه بکاپ‌ها دسترسی پیدا کند
    safe_name = secure_filename(filename)
    if not safe_name or safe_name != filename:
        return jsonify({"ok": False, "message": "نام فایل پشتیبان نامعتبر است"}), 400
    src = os.path.realpath(os.path.join(BACKUP_DIR, safe_name))
    backup_dir_real = os.path.realpath(BACKUP_DIR)
    if os.path.commonpath([src, backup_dir_real]) != backup_dir_real:
        return jsonify({"ok": False, "message": "نام فایل پشتیبان نامعتبر است"}), 400
    if not os.path.exists(src):
        return jsonify({"ok": False, "message": "فایل پشتیبان پیدا نشد"}), 404

    do_backup()  # نسخه احتیاطی از وضعیت فعلی قبل از بازنویسی
    shutil.copy2(src, DB_PATH)
    log_action(d.get("username"), "بازیابی بکاپ", filename)
    return jsonify({"ok": True, "message": f'دیتابیس از نسخه «{filename}» بازیابی شد. سرور و کلاینت را ببندید و دوباره باز کنید.'})


@app.route("/reports/debtors", methods=["GET"])
def report_debtors():
    """لیست مشتریانی که به ما بدهکارند"""
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM parties WHERE type='customer' AND balance > 0 ORDER BY balance DESC"
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/reports/creditors", methods=["GET"])
def report_creditors():
    """لیست تامین‌کنندگانی که ما به آن‌ها بدهکاریم (بستانکاران ما)"""
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM parties WHERE type='supplier' AND balance < 0 ORDER BY balance ASC"
    ).fetchall()
    conn.close()
    result = [dict(r) for r in rows]
    for r in result:
        r["owed_amount"] = abs(r["balance"])
    return jsonify(result)


@app.route("/reports/top-items", methods=["GET"])
def report_top_items():
    """پرفروش‌ترین کالاها در N روز اخیر (پیش‌فرض ۳۰ روز)"""
    days = request.args.get("days", default=30, type=int)
    conn = get_connection()
    since = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    rows = conn.execute("""
        SELECT items.id as item_id, items.name as name, items.brand as brand,
               SUM(invoice_items.qty) as total_qty, SUM(invoice_items.total) as total_amount
        FROM invoice_items
        JOIN invoices ON invoice_items.invoice_id = invoices.id
        JOIN items ON invoice_items.item_id = items.id
        WHERE invoices.invoice_type='sale' AND invoices.voided=0 AND invoices.date >= ?
        GROUP BY items.id
        ORDER BY total_amount DESC
        LIMIT 20
    """, (since,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/reports/profit-by-item", methods=["GET"])
def report_profit_by_item():
    """سود تقریبی هر کالا در N روز اخیر — بر اساس قیمت فروش هر ردیف فاکتور منهای بهای
    تمام‌شده‌ی فعلی کالا (میانگین موزون قیمت خرید، یا اگر نبود، آخرین قیمت خرید). فقط برای مدیر"""
    err = require_admin()
    if err:
        return err
    days = request.args.get("days", default=30, type=int)
    conn = get_connection()
    since = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    rows = conn.execute("""
        SELECT items.id as item_id, items.name as name, items.brand as brand,
               COALESCE(items.avg_cost, items.purchase_price, 0) as unit_cost,
               SUM(invoice_items.qty) as total_qty, SUM(invoice_items.total) as total_sales
        FROM invoice_items
        JOIN invoices ON invoice_items.invoice_id = invoices.id
        JOIN items ON invoice_items.item_id = items.id
        WHERE invoices.invoice_type='sale' AND invoices.voided=0 AND invoices.date >= ?
        GROUP BY items.id
        ORDER BY total_sales DESC
    """, (since,)).fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        d["estimated_cost"] = d["unit_cost"] * d["total_qty"]
        d["estimated_profit"] = d["total_sales"] - d["estimated_cost"]
        result.append(d)
    return jsonify(result)


@app.route("/reports/reorder-suggestions", methods=["GET"])
def report_reorder_suggestions():
    """پیشنهاد سفارش مجدد: بر اساس میانگین فروش روزانه‌ی هر کالا در ۳۰ روز اخیر،
    تخمین می‌زند چند روز دیگر موجودی تمام می‌شود و چقدر باید سفارش داد."""
    conn = get_connection()
    since = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    rows = conn.execute("""
        SELECT items.id as item_id, items.name as name, items.brand as brand,
               items.stock_qty as stock_qty, items.min_stock as min_stock, items.unit as unit,
               COALESCE(SUM(invoice_items.qty), 0) as sold_30d
        FROM items
        LEFT JOIN invoice_items ON invoice_items.item_id = items.id
        LEFT JOIN invoices ON invoice_items.invoice_id = invoices.id
            AND invoices.invoice_type='sale' AND invoices.voided=0 AND invoices.date >= ?
        WHERE items.deleted_at IS NULL
        GROUP BY items.id
    """, (since,)).fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        daily_rate = d["sold_30d"] / 30.0
        if daily_rate <= 0:
            continue
        days_left = d["stock_qty"] / daily_rate if daily_rate > 0 else None
        if days_left is None or days_left > 14:
            continue
        suggested_qty = max((daily_rate * 30) - d["stock_qty"], 0)
        d["daily_rate"] = round(daily_rate, 2)
        d["days_left"] = round(days_left, 1)
        d["suggested_reorder_qty"] = round(suggested_qty, 2)
        result.append(d)
    result.sort(key=lambda x: x["days_left"])
    return jsonify(result)


@app.route("/reports/yoy-comparison", methods=["GET"])
def report_yoy_comparison():
    """مقایسه فروش این ماه و امسال با همان بازه در سال قبل (تقویم میلادی، چون تاریخ‌ها
    در دیتابیس میلادی ذخیره می‌شوند؛ فقط برای مقایسه‌ی نسبی کاربردی است)"""
    conn = get_connection()
    now_dt = datetime.now()

    def sales_since(since_dt, until_dt=None):
        q = "SELECT COALESCE(SUM(total),0) as t FROM invoices WHERE invoice_type='sale' AND voided=0 AND date >= ?"
        params = [since_dt.strftime("%Y-%m-%d")]
        if until_dt:
            q += " AND date < ?"
            params.append(until_dt.strftime("%Y-%m-%d"))
        return conn.execute(q, params).fetchone()["t"]

    this_month_start = now_dt.replace(day=1)
    last_month_year_start = this_month_start.replace(year=this_month_start.year - 1)
    try:
        last_month_year_end = last_month_year_start.replace(month=last_month_year_start.month % 12 + 1) \
            if last_month_year_start.month != 12 else last_month_year_start.replace(year=last_month_year_start.year + 1, month=1)
    except ValueError:
        last_month_year_end = last_month_year_start + timedelta(days=28)

    this_year_start = now_dt.replace(month=1, day=1)
    last_year_start = this_year_start.replace(year=this_year_start.year - 1)
    last_year_end = this_year_start

    result = {
        "this_month": sales_since(this_month_start),
        "same_month_last_year": sales_since(last_month_year_start, last_month_year_end),
        "this_year": sales_since(this_year_start),
        "last_year": sales_since(last_year_start, last_year_end),
    }
    conn.close()
    return jsonify(result)


@app.route("/reports/by-employee", methods=["GET"])
def report_by_employee():
    """فروش هر کارمند در N روز اخیر (پیش‌فرض ۳۰ روز) — فقط برای مدیر"""
    err = require_admin()
    if err:
        return err
    days = request.args.get("days", default=30, type=int)
    conn = get_connection()
    since = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    rows = conn.execute("""
        SELECT COALESCE(created_by, 'نامشخص') as username,
               COUNT(*) as invoice_count, SUM(total) as total_amount
        FROM invoices
        WHERE invoice_type='sale' AND voided=0 AND date >= ?
        GROUP BY username
        ORDER BY total_amount DESC
    """, (since,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/invoices/<int:invoice_id>/print", methods=["GET"])
def invoice_print(invoice_id):
    """
    فاکتور را به‌صورت یک صفحه HTML آماده چاپ برمی‌گرداند (نه PDF).
    چون توسط مرورگر رندر می‌شود، متن فارسی همیشه کاملاً درست نمایش داده می‌شود.
    """
    page_format = request.args.get("format", "A5")
    conn = get_connection()
    invoice = conn.execute(
        "SELECT invoices.* FROM invoices WHERE invoices.id=?", (invoice_id,)
    ).fetchone()
    if not invoice:
        conn.close()
        return "فاکتور پیدا نشد", 404
    items = conn.execute("""
        SELECT invoice_items.qty, invoice_items.unit_price, invoice_items.total,
               invoice_items.serial_number, invoice_items.warranty_months,
               items.name as item_name, items.brand as brand, categories.name as category_name
        FROM invoice_items
        JOIN items ON invoice_items.item_id = items.id
        LEFT JOIN categories ON items.category_id = categories.id
        WHERE invoice_id=?
    """, (invoice_id,)).fetchall()
    party = None
    if invoice["party_id"]:
        party = conn.execute("SELECT * FROM parties WHERE id=?", (invoice["party_id"],)).fetchone()
    conn.close()

    cfg = load_shop_settings()

    inv_dict = dict(invoice)
    items_list = [dict(r) for r in items]
    party_dict = dict(party) if party else None
    words_text = pdf_generator.number_to_persian_words(inv_dict["total"])

    html = invoice_html.build_invoice_html(
        inv_dict, items_list, party_dict, page_format=page_format,
        shop_name=cfg.get("name") or "حسابداری",
        shop_phones=cfg.get("phones", ""),
        shop_address=cfg.get("address", ""),
        words_text=words_text,
        logo_url=(f"/assets/{cfg['logo_filename']}" if cfg.get("logo_filename") else None),
        footer_message=cfg.get("invoice_footer_message", ""),
        shop_national_id=cfg.get("national_id", ""),
        shop_economic_code=cfg.get("economic_code", ""),
        shop_postal_code=cfg.get("postal_code", ""),
    )
    return html


@app.route("/invoices/<int:invoice_id>/pdf", methods=["GET"])
def invoice_pdf(invoice_id):
    page_format = request.args.get("format", "A5")
    conn = get_connection()
    invoice = conn.execute(
        "SELECT invoices.*, parties.name as party_name FROM invoices "
        "LEFT JOIN parties ON invoices.party_id = parties.id WHERE invoices.id=?", (invoice_id,)
    ).fetchone()
    if not invoice:
        conn.close()
        return jsonify({"ok": False, "message": "فاکتور پیدا نشد"}), 404
    items = conn.execute(
        "SELECT invoice_items.*, items.name as item_name FROM invoice_items "
        "JOIN items ON invoice_items.item_id = items.id WHERE invoice_id=?", (invoice_id,)
    ).fetchall()
    conn.close()

    inv_dict = dict(invoice)
    inv_dict["id"] = inv_dict.get("number") or inv_dict["id"]
    items_list = [dict(r) for r in items]
    tmp_path = os.path.join(tempfile.gettempdir(), f"invoice_{invoice_id}_{page_format}.pdf")
    pdf_generator.generate_invoice_pdf(
        tmp_path, inv_dict, items_list, party_name=invoice["party_name"], page_format=page_format
    )
    return send_file(tmp_path, mimetype="application/pdf")


@app.route("/parties/<int:party_id>/statement/print", methods=["GET"])
def party_statement_print(party_id):
    """صورت‌حساب کامل یک طرف‌حساب به‌صورت HTML آماده چاپ (مثل چاپ فاکتور تکی)"""
    conn = get_connection()
    party = conn.execute("SELECT * FROM parties WHERE id=?", (party_id,)).fetchone()
    if not party:
        conn.close()
        return "طرف حساب پیدا نشد", 404
    invoices = conn.execute("SELECT * FROM invoices WHERE party_id=? AND voided=0 ORDER BY date", (party_id,)).fetchall()
    conn.close()

    cfg = load_shop_settings()
    html = invoice_html.build_statement_html(
        dict(party), [dict(r) for r in invoices],
        shop_name=cfg.get("name") or "حسابداری",
        shop_phones=cfg.get("phones", ""),
        shop_address=cfg.get("address", ""),
        logo_url=(f"/assets/{cfg['logo_filename']}" if cfg.get("logo_filename") else None),
    )
    return html


@app.route("/parties/<int:party_id>/statement/pdf", methods=["GET"])
def party_statement_pdf(party_id):
    conn = get_connection()
    party = conn.execute("SELECT * FROM parties WHERE id=?", (party_id,)).fetchone()
    if not party:
        conn.close()
        return jsonify({"ok": False, "message": "طرف حساب پیدا نشد"}), 404
    invoices = conn.execute("SELECT * FROM invoices WHERE party_id=? AND voided=0 ORDER BY date", (party_id,)).fetchall()
    conn.close()

    cfg = load_shop_settings()
    tmp_path = os.path.join(tempfile.gettempdir(), f"statement_{party_id}.pdf")
    pdf_generator.generate_statement_pdf(
        tmp_path, dict(party), [dict(r) for r in invoices], shop_name=cfg.get("name") or "حسابداری"
    )
    return send_file(tmp_path, mimetype="application/pdf")


@app.route("/export/items.xlsx", methods=["GET"])
def export_items_excel():
    import openpyxl
    conn = get_connection()
    rows = conn.execute("SELECT * FROM items WHERE deleted_at IS NULL ORDER BY name").fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "کالاها"
    ws.append(["کد/بارکد", "نام", "برند", "واحد", "قیمت خرید", "قیمت فروش", "موجودی", "حداقل موجودی"])
    for r in rows:
        ws.append([r["code"] or "", r["name"], r["brand"] or "", r["unit"],
                   r["purchase_price"], r["sale_price"], r["stock_qty"], r["min_stock"]])
    tmp_path = os.path.join(tempfile.gettempdir(), "items_export.xlsx")
    wb.save(tmp_path)
    return send_file(tmp_path, as_attachment=True, download_name="کالاها.xlsx")


@app.route("/import/items.xlsx", methods=["POST"])
def import_items_excel():
    """ورودی گروهی کالا از فایل اکسل — همون فرمت خروجی اکسل (کد/بارکد، نام، برند، واحد،
    قیمت خرید، قیمت فروش، موجودی، حداقل موجودی). ردیف‌های خراب رد می‌شوند، بقیه ثبت می‌شوند."""
    err = require_permission("can_manage_items")
    if err:
        return err
    if "file" not in request.files:
        return jsonify({"ok": False, "message": "فایلی ارسال نشده"}), 400
    file = request.files["file"]
    if not file or file.filename == "":
        return jsonify({"ok": False, "message": "فایلی انتخاب نشده"}), 400
    if not file.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "message": "فایل باید با فرمت xlsx باشد"}), 400

    import openpyxl
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        return jsonify({"ok": False, "message": "فایل اکسل قابل خواندن نبود"}), 400

    conn = get_connection()
    imported_count = 0
    errors = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        code, name, brand, unit, purchase_price, sale_price, stock_qty, min_stock = (list(row) + [None] * 8)[:8]
        name = str(name).strip() if name is not None else ""
        if not name:
            errors.append(f"ردیف {row_num}: نام کالا خالی است")
            continue
        try:
            purchase_price = float(purchase_price) if purchase_price not in (None, "") else 0
            sale_price = float(sale_price) if sale_price not in (None, "") else 0
            stock_qty = float(stock_qty) if stock_qty not in (None, "") else 0
            min_stock = float(min_stock) if min_stock not in (None, "") else 0
        except (TypeError, ValueError):
            errors.append(f"ردیف {row_num} ({name}): قیمت یا موجودی عدد معتبر نیست")
            continue
        try:
            conn.execute(
                """INSERT INTO items (code, name, category_id, unit, purchase_price, sale_price, stock_qty, min_stock, brand)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (str(code).strip() if code else None, name, None, str(unit).strip() if unit else "عدد",
                 purchase_price, sale_price, stock_qty, min_stock, str(brand).strip() if brand else None)
            )
            imported_count += 1
        except sqlite3.IntegrityError:
            errors.append(f"ردیف {row_num} ({name}): کد/بارکد تکراری است")
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "imported_count": imported_count, "skipped_count": len(errors), "errors": errors[:30]})


@app.route("/items/labels/print", methods=["GET"])
def print_item_labels():
    """برگه برچسب قیمت/بارکد آماده چاپ برای کالاهای انتخاب‌شده (?ids=1,2,3)"""
    ids_param = request.args.get("ids", "")
    try:
        item_ids = [int(x) for x in ids_param.split(",") if x.strip()]
    except ValueError:
        return "شناسه کالا نامعتبر است", 400
    if not item_ids:
        return "کالایی انتخاب نشده", 400

    conn = get_connection()
    placeholders = ",".join("?" * len(item_ids))
    rows = conn.execute(f"SELECT * FROM items WHERE id IN ({placeholders})", item_ids).fetchall()
    conn.close()
    items_by_id = {r["id"]: dict(r) for r in rows}
    ordered_items = [items_by_id[i] for i in item_ids if i in items_by_id]
    return invoice_html.build_labels_html(ordered_items)


@app.route("/ai/analyze-invoice-photo", methods=["POST"])
def analyze_invoice_photo():
    """آنالیز عکس فاکتور خرید با Claude (نیاز به اینترنت و کلید API در config.json دارد)"""
    if "photo" not in request.files:
        return jsonify({"ok": False, "message": "عکسی ارسال نشده"}), 400
    file = request.files["photo"]
    if not file or file.filename == "":
        return jsonify({"ok": False, "message": "عکسی انتخاب نشده"}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    media_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png", ".webp": "image/webp", ".pdf": "application/pdf"}
    media_type = media_map.get(ext)
    if not media_type:
        return jsonify({"ok": False, "message": "فرمت فایل باید jpg، png، webp یا pdf باشد"}), 400

    method = request.form.get("method", "ai")
    image_bytes = file.read()

    if method == "offline":
        ok, data = invoice_ocr_free.analyze_invoice_image_offline(image_bytes, is_pdf=(ext == ".pdf"))
        if not ok:
            return jsonify({"ok": False, "message": data}), 400
        return jsonify({"ok": True, "data": data, "method": "offline"})

    s = load_shop_settings()
    if not s.get("ai_enabled") or not s.get("ai_api_key"):
        return jsonify({"ok": False, "message": "روش هوش مصنوعی فعال نشده. از «تنظیمات کلی › دستیار هوش مصنوعی» کلید API را وارد کن، یا روش «رایگان آفلاین» را انتخاب کن"}), 400

    ok, data = invoice_ai.analyze_invoice_image(
        image_bytes, media_type, s.get("ai_api_key"), "claude-haiku-4-5-20251001"
    )
    if not ok:
        return jsonify({"ok": False, "message": data}), 400
    return jsonify({"ok": True, "data": data, "method": "ai"})


@app.route("/assistant/ask", methods=["POST"])
def assistant_ask():
    """دستیار هوشمند برنامه — فقط درباره‌ی نحوه‌ی کار با خود برنامه پاسخ می‌دهد
    (نیاز به اینترنت و همان کلید API که برای اسکن فاکتور خرید در config.json تنظیم شده دارد)"""
    d = request.json or {}
    question = d.get("question", "")
    history = d.get("history", [])

    s = load_shop_settings()
    if not s.get("ai_enabled") or not s.get("ai_api_key"):
        return jsonify({"ok": False, "message": "دستیار هوشمند فعال نیست. مدیر باید از «تنظیمات کلی › دستیار هوش مصنوعی» کلید API را وارد کند."}), 400

    ok, answer = assistant_ai.ask_assistant(question, history, s.get("ai_api_key"), "claude-haiku-4-5-20251001")
    if not ok:
        return jsonify({"ok": False, "message": answer}), 400
    return jsonify({"ok": True, "answer": answer})


@app.route("/items/apply-bulk-prices", methods=["POST"])
def apply_bulk_prices():
    """
    اعمال گروهی قیمت جدید روی چند کالا، با ثبت تاریخچه قبل از تغییر.
    ورودی: {"changes": [{"item_id": 1, "new_purchase_price": 5500, "new_sale_price": 8800}, ...], "note": "...", "username": "..."}
    """
    err = require_admin()
    if err:
        return err
    d = request.json
    changes = d.get("changes", [])
    if not changes:
        return jsonify({"ok": False, "message": "هیچ تغییری ارسال نشده"}), 400

    conn = get_connection()
    count = 0
    for ch in changes:
        item = conn.execute("SELECT * FROM items WHERE id=?", (ch["item_id"],)).fetchone()
        if not item:
            continue
        conn.execute(
            "INSERT INTO price_history (item_id, old_purchase_price, old_sale_price, new_purchase_price, new_sale_price, changed_at, note, username) "
            "VALUES (?,?,?,?,?,?,?,?)",
            (ch["item_id"], item["purchase_price"], item["sale_price"],
             ch["new_purchase_price"], ch["new_sale_price"], now(), d.get("note", ""), d.get("username"))
        )
        conn.execute("UPDATE items SET purchase_price=?, sale_price=? WHERE id=?",
                     (ch["new_purchase_price"], ch["new_sale_price"], ch["item_id"]))
        count += 1
    conn.commit()
    conn.close()
    log_action(d.get("username"), "افزایش/تغییر گروهی قیمت", f"{count} کالا تغییر کرد — {d.get('note', '')}")
    return jsonify({"ok": True, "updated_count": count})


@app.route("/price-history", methods=["GET"])
def get_price_history():
    err = require_admin()
    if err:
        return err
    conn = get_connection()
    rows = conn.execute("""
        SELECT price_history.*, items.name as item_name
        FROM price_history JOIN items ON price_history.item_id = items.id
        ORDER BY price_history.id DESC LIMIT 500
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/price-history/<int:history_id>/revert", methods=["POST"])
def revert_price_history(history_id):
    """بازگردانی قیمت یک کالا به مقدار قبل از یک تغییر خاص (خودش هم یک رکورد تاریخچه جدید می‌سازد)"""
    err = require_admin()
    if err:
        return err
    d = request.json or {}
    conn = get_connection()
    h = conn.execute("SELECT * FROM price_history WHERE id=?", (history_id,)).fetchone()
    if not h:
        conn.close()
        return jsonify({"ok": False, "message": "رکورد تاریخچه پیدا نشد"}), 404
    item = conn.execute("SELECT * FROM items WHERE id=?", (h["item_id"],)).fetchone()
    if not item:
        conn.close()
        return jsonify({"ok": False, "message": "کالا پیدا نشد"}), 404

    conn.execute(
        "INSERT INTO price_history (item_id, old_purchase_price, old_sale_price, new_purchase_price, new_sale_price, changed_at, note, username) "
        "VALUES (?,?,?,?,?,?,?,?)",
        (h["item_id"], item["purchase_price"], item["sale_price"],
         h["old_purchase_price"], h["old_sale_price"], now(), "بازگردانی به نسخه قبلی", d.get("username"))
    )
    conn.execute("UPDATE items SET purchase_price=?, sale_price=? WHERE id=?",
                 (h["old_purchase_price"], h["old_sale_price"], h["item_id"]))
    conn.commit()
    conn.close()
    log_action(d.get("username"), "بازگردانی قیمت", f'کالا: {item["name"]}')
    return jsonify({"ok": True})


# ---------- ویجت‌های تکمیلی داشبورد ----------
@app.route("/reports/stock-ranking", methods=["GET"])
def stock_ranking():
    """رتبه‌بندی کالاها بر اساس موجودی فعلی (بیشترین تا کمترین)"""
    conn = get_connection()
    rows = conn.execute("SELECT name, brand, stock_qty, unit FROM items WHERE deleted_at IS NULL ORDER BY stock_qty DESC LIMIT 10").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/ping", methods=["GET"])
def ping():
    return jsonify({"ok": True, "message": "سرور فعال است"})


def run_embedded(port=5050):
    """
    اجرای این سرور به‌صورت تعبیه‌شده داخل خود برنامه دسکتاپ (فقط روی 127.0.0.1، تک‌کاربره).
    این تابع باید در یک ترد جدا (daemon thread) فراخوانی شود.
    """
    import logging
    logging.getLogger("werkzeug").setLevel(logging.ERROR)  # کم کردن پیام‌های اضافه در کنسول
    init_db()
    threading.Thread(target=backup_loop, daemon=True).start()
    threading.Thread(target=nightly_loop, daemon=True).start()
    threading.Thread(target=weekly_recap_loop, daemon=True).start()
    app.run(host="127.0.0.1", port=port, threaded=True, use_reloader=False, debug=False)


if __name__ == "__main__":
    init_db()
    threading.Thread(target=backup_loop, daemon=True).start()
    threading.Thread(target=nightly_loop, daemon=True).start()
    threading.Thread(target=weekly_recap_loop, daemon=True).start()
    print("=" * 50)
    print("سرور حسابداری در حال اجراست...")
    print("این پنجره را باز نگه دارید تا کلاینت‌ها بتوانند وصل شوند.")
    print("نسخه‌های پشتیبان به‌صورت خودکار در پوشه backups ذخیره می‌شوند.")
    print("گزارش شبانه (ایمیل + آپلود VPS) هر شب ساعت ۲۳ اجرا می‌شود (در صورت تنظیم config.json).")
    print("=" * 50)
    # روی همه رابط‌های شبکه گوش می‌ده تا از کامپیوترهای دیگه قابل دسترسی باشه
    app.run(host="0.0.0.0", port=5050, debug=False)
